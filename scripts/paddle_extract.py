# WORKING FOR 2025 AND 2024

# # -*- coding: utf-8 -*-
# """
# nirf_extract.py  -  NIRF score-card image -> Excel extractor
# Install: pip install pytesseract openpyxl Pillow
# Tesseract binary: https://github.com/UB-Mannheim/tesseract/wiki
# Usage: python scripts/paddle_extract.py <image_or_folder> [--output out.xlsx] [--workers N]
# """

# import io, os, sys, re, json, argparse
# from pathlib import Path
# from concurrent.futures import ThreadPoolExecutor, as_completed
# import multiprocessing

# if hasattr(sys.stdout, "reconfigure"):
#     sys.stdout.reconfigure(encoding="utf-8", errors="replace")
# else:
#     sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# # TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
# TESSERACT_CMD = None


# def setup_tesseract():
#     import pytesseract
#     if TESSERACT_CMD:
#         pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
#     else:
#         default = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
#         if os.path.exists(default):
#             pytesseract.pytesseract.tesseract_cmd = default
#     return pytesseract


# def open_image(path):
#     from PIL import Image
#     return Image.open(path).convert("RGB")

# def to_numpy(img):
#     import numpy as np
#     return np.array(img)


# # ── ORIGINAL border detection (was working) ───────────────────────────────────

# def find_borders(arr):
#     import numpy as np
#     h, w = arr.shape[:2]
#     threshold = w * 0.3
#     borders = []
#     in_border = False
#     for y in range(h):
#         row = arr[y]
#         dark = int(np.sum((row[:,0] < 100) & (row[:,1] < 100) & (row[:,2] < 100)))
#         if dark > threshold:
#             if not in_border:
#                 borders.append(y)
#                 in_border = True
#         else:
#             in_border = False
#     return borders


# # ── ORIGINAL OCR helpers (were working) ──────────────────────────────────────

# def ocr_band_text(pytesseract, band_img, digits_only=False):
#     from PIL import Image
#     import numpy as np
#     w, h = band_img.size
#     band_img = band_img.resize((w * 3, h * 3), Image.LANCZOS)
#     gray = band_img.convert("L")
#     arr = np.array(gray)
#     if arr.mean() < 128:
#         gray = gray.point(lambda x: 255 - x)
#     config = "--psm 7 -c tessedit_char_whitelist=" + (
#         "0123456789. " if digits_only else
#         "ABCDEFGHIJKLMNOPQRSTUVWXYZ+. 0123456789"
#     )
#     return pytesseract.image_to_string(gray, config=config).strip()


# def ocr_band_data(pytesseract, band_img):
#     from PIL import Image
#     import numpy as np
#     w, h = band_img.size
#     band_img = band_img.resize((w * 3, h * 3), Image.LANCZOS)
#     gray = band_img.convert("L")
#     arr = np.array(gray)
#     if arr.mean() < 128:
#         gray = gray.point(lambda x: 255 - x)
#     config = "--psm 11 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ+.0123456789"
#     data = pytesseract.image_to_data(
#         gray, config=config, output_type=pytesseract.Output.DICT
#     )
#     items = []
#     for i, text in enumerate(data["text"]):
#         text = text.strip()
#         if not text or data["conf"][i] < 10:
#             continue
#         x_center = (data["left"][i] + data["width"][i] / 2) / 3
#         items.append((x_center, text))
#     items.sort(key=lambda t: t[0])
#     return items


# GROUP_HEADERS = {"TLR", "RP", "GO", "OI"}
# ROW_LABELS    = {"SCORE", "TOTAL", "RANK"}

# # Known NIRF column headers — used to correct common OCR misreads
# KNOWN_COLS = {
#     "SS", "FSR", "FQE", "FRU", "OE+MIR", "OE",
#     "PU", "QP", "IPR", "FPPP", "SDG",
#     "GUE", "GPHD",
#     "RD", "WD", "ESCS", "PCS",
#     "PR", "MS", "GPH", "PERCEPTION",
# }

# # Maps common tesseract misreads -> correct token
# OCR_FIXES = {
#     "S5": "SS", "5S": "SS", "S$": "SS", "55": "SS",
#     "W0": "WD", "VVD": "WD", "WO": "WD", "W0": "WD",
#     "FQF": "FQE", "FRO": "FRU", "FR0": "FRU",
#     "1PR": "IPR", "lPR": "IPR",
#     "GU3": "GUE",
#     "P5": "PCS", "PC5": "PCS",
#     "E5CS": "ESCS", "ESC5": "ESCS",
#     "0E+MIR": "OE+MIR", "OE-MIR": "OE+MIR",
#     "5DG": "SDG", "SDC": "SDG", "5DC": "SDG",
# }

# def fix_ocr(tok):
#     """Correct common tesseract misreads for NIRF column headers."""
#     t = tok.upper().replace(" ", "")
#     if t in OCR_FIXES:
#         return OCR_FIXES[t]
#     return t

# def is_float(s):
#     try:
#         float(s)
#         return True
#     except ValueError:
#         return False

# def is_col_header(tok):
#     t = fix_ocr(tok)
#     if t in GROUP_HEADERS or t in ROW_LABELS:
#         return False
#     # Accept known cols directly
#     if t in KNOWN_COLS:
#         return True
#     # Also accept any 2-8 uppercase letter token
#     return bool(re.fullmatch(r"[A-Z+]{2,8}", t))


# def match_by_x(header_items, value_items):
#     result = {}
#     used = set()
#     for hx, htext in header_items:
#         best_dist, best_idx = float("inf"), None
#         for i, (vx, _) in enumerate(value_items):
#             if i in used:
#                 continue
#             dist = abs(hx - vx)
#             if dist < best_dist:
#                 best_dist, best_idx = dist, i
#         if best_idx is not None:
#             used.add(best_idx)
#             result[htext] = value_items[best_idx][1]
#     return result


# def extract_one(image_path):
#     pytesseract = setup_tesseract()
#     import numpy as np

#     img = open_image(image_path)
#     arr = to_numpy(img)
#     h, w = arr.shape[:2]

#     # ── Institute code from filename (always correct) ─────────────────────────
#     fname = Path(image_path).stem
#     institute_code = fname if re.match(r"IR-[A-Z]-[A-Z]-\d+", fname) else ""

#     # ── Institute name: find title row by scanning top 80px for non-white ─────
#     title_start = title_end = None
#     for y in range(min(80, h)):
#         non_white = int(np.sum(~((arr[y,:,0] > 220) & (arr[y,:,1] > 220) & (arr[y,:,2] > 220))))
#         if non_white > 20:
#             if title_start is None:
#                 title_start = y
#             title_end = y
#     if title_start is None:
#         title_start, title_end = 15, 32

#     # Crop with padding and upscale for tesseract
#     t1, t2 = max(0, title_start - 4), min(h, title_end + 4)
#     title_crop = img.crop((0, t1, w, t2))
#     tw, th = title_crop.size
#     title_big = title_crop.resize((tw * 4, th * 4), __import__('PIL').Image.LANCZOS)
#     gray = title_big.convert("L")
#     # PSM 7 = single line, no whitelist so lowercase is allowed
#     title_text = pytesseract.image_to_string(gray, config="--psm 7").strip()

#     # Extract name = text before "(IR-...)"
#     m = re.search(r"(.+?)\s*\(IR-[A-Z]-[A-Z]-\d+\)", title_text)
#     if m:
#         institute_name = m.group(1).strip()
#     else:
#         lines = [l.strip() for l in title_text.splitlines() if len(l.strip()) > 4]
#         institute_name = max(lines, key=len) if lines else ""

#     # ── ORIGINAL table extraction (unchanged) ─────────────────────────────────
#     borders = find_borders(arr)
#     if len(borders) < 4:
#         borders = [int(h*0.357), int(h*0.390), int(h*0.425), int(h*0.459)]

#     b0, b1, b2, b3 = borders[0], borders[1], borders[2], borders[3]
#     pad = 3

#     col_img   = img.crop((0, b1+pad, w, b2-pad))
#     score_img = img.crop((0, b2+pad, w, b3-pad))
#     total_img = img.crop((0, b3+pad, w, min(b3+48, h)))

#     col_items   = ocr_band_data(pytesseract, col_img)
#     score_items = ocr_band_data(pytesseract, score_img)
#     total_items = ocr_band_data(pytesseract, total_img)

#     header_items = [(x, fix_ocr(t)) for x, t in col_items if is_col_header(t)]
#     score_vals   = [(x, t)         for x, t in score_items if is_float(t)]
#     total_vals   = [(x, t)         for x, t in total_items if is_float(t)]

#     score_map = match_by_x(header_items, score_vals)
#     total_map = match_by_x(header_items, total_vals)

#     table = {
#         hname: {"score": score_map.get(hname, ""), "total": total_map.get(hname, "")}
#         for _, hname in header_items
#     }

#     parts = Path(image_path).resolve().parts
#     year = category = ""
#     for i, p in enumerate(parts):
#         if p == "image" and i >= 2:
#             year, category = parts[i-2], parts[i-1]
#             break

#     n_cols   = len(table)
#     n_scores = sum(1 for v in table.values() if v["score"])
#     flag     = " MISMATCH" if n_cols != n_scores and n_cols > 0 else " OK"

#     return {
#         "image_path":     image_path,
#         "year":           year,
#         "category":       category,
#         "institute_name": institute_name,
#         "institute_code": institute_code,
#         "table":          table,
#         "_msg":           f"  {institute_code} | {institute_name[:40] if institute_name else '?'} | cols={n_cols} scores={n_scores}{flag}",
#     }


# IMAGE_EXTS = {".jpg", ".jpeg", ".png"}

# def collect_images(path):
#     p = Path(os.path.normpath(path))
#     if p.is_file() and p.suffix.lower() in IMAGE_EXTS:
#         return [str(p)]
#     if p.is_dir():
#         imgs = []
#         for ext in IMAGE_EXTS:
#             imgs.extend(str(f) for f in p.rglob(f"*{ext}"))
#         imgs.sort()
#         return imgs
#     print(f"Path not found: {p}", file=sys.stderr)
#     return []


# def build_excel(rows, output_path):
#     from openpyxl import Workbook
#     from openpyxl.styles import Font, PatternFill, Alignment

#     header_fill  = PatternFill("solid", fgColor="1F4E79")
#     header_font  = Font(color="FFFFFF", bold=True)
#     alt_fill     = PatternFill("solid", fgColor="D9E1F2")
#     center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

#     all_headers, seen_h = [], set()
#     for row in rows:
#         for hk in row.get("table", {}):
#             if hk not in seen_h:
#                 seen_h.add(hk)
#                 all_headers.append(hk)

#     fixed_cols = ["Year", "Category", "Institute Name", "Institute Code"]
#     all_cols   = fixed_cols \
#                + [f"{h} Score" for h in all_headers] \
#                + [f"{h} Total" for h in all_headers]

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "NIRF Scores"

#     for ci, col in enumerate(all_cols, 1):
#         cell = ws.cell(row=1, column=ci, value=col)
#         cell.fill = header_fill
#         cell.font = header_font
#         cell.alignment = center_align
#     ws.row_dimensions[1].height = 30

#     for ri, row in enumerate(rows, 2):
#         tbl  = row.get("table", {})
#         fill = alt_fill if ri % 2 == 0 else PatternFill()
#         vals = [
#             row.get("year", ""),
#             row.get("category", ""),
#             row.get("institute_name", ""),
#             row.get("institute_code", ""),
#         ] + [tbl.get(h, {}).get("score", "") for h in all_headers] \
#           + [tbl.get(h, {}).get("total", "") for h in all_headers]

#         for ci, val in enumerate(vals, 1):
#             cell = ws.cell(row=ri, column=ci, value=val)
#             cell.fill = fill
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#     for ci in range(1, len(all_cols) + 1):
#         col_letter = ws.cell(row=1, column=ci).column_letter
#         max_len = max(
#             len(str(ws.cell(row=1, column=ci).value or "")),
#             *(len(str(ws.cell(row=ri, column=ci).value or "")) for ri in range(2, len(rows)+2))
#         )
#         ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

#     ws.freeze_panes = "A2"
#     wb.save(output_path)
#     print(f"\nDone. Saved {len(rows)} rows -> {output_path}", flush=True)


# def main():
#     parser = argparse.ArgumentParser(description="Extract NIRF score-card table data")
#     parser.add_argument("input",     help="Image file or folder (recursive)")
#     parser.add_argument("--output",  "-o", default="nirf_scores.xlsx")
#     parser.add_argument("--json",    "-j", default=None)
#     parser.add_argument("--workers", "-w", type=int,
#                         default=min(8, multiprocessing.cpu_count()),
#                         help="Parallel threads (default: min(8, cpu_count))")
#     args = parser.parse_args()

#     images = collect_images(args.input)
#     if not images:
#         print(f"No images found at: {args.input}", file=sys.stderr)
#         sys.exit(1)

#     n = len(images)
#     w = min(args.workers, n)
#     print(f"Found {n} image(s). Running with {w} thread(s)...\n", flush=True)

#     try:
#         pyt = setup_tesseract()
#         from PIL import Image as PILImage
#         import numpy as np
#         test = PILImage.fromarray(np.ones((10,10,3), dtype=np.uint8)*255)
#         pyt.image_to_string(test)
#         print("Tesseract ready.\n", flush=True)
#     except Exception as e:
#         print(f"\nERROR: Tesseract not found: {e}", file=sys.stderr)
#         print("Install from: https://github.com/UB-Mannheim/tesseract/wiki", file=sys.stderr)
#         print("Then add to PATH or set TESSERACT_CMD at top of this script.", file=sys.stderr)
#         sys.exit(1)

#     rows_by_path = {}
#     done = 0

#     with ThreadPoolExecutor(max_workers=w) as exe:
#         futures = {exe.submit(extract_one, p): p for p in images}
#         for fut in as_completed(futures):
#             done += 1
#             img_path = futures[fut]
#             try:
#                 row = fut.result()
#                 msg = row.pop("_msg", "")
#                 rows_by_path[img_path] = row
#                 print(f"[{done}/{n}] {Path(img_path).name}{msg}", flush=True)
#             except Exception as e:
#                 print(f"[{done}/{n}] ERROR {Path(img_path).name}: {e}", flush=True)

#     rows = [rows_by_path[p] for p in images if p in rows_by_path]

#     if args.json:
#         with open(args.json, "w", encoding="utf-8") as f:
#             json.dump(rows, f, indent=2, ensure_ascii=False)
#         print(f"JSON -> {args.json}", flush=True)

#     if not rows:
#         print("No rows extracted.", flush=True)
#         sys.exit(0)

#     build_excel(rows, args.output)


# if __name__ == "__main__":
#     main()

















#LATEST WORKING FOR 2023



# # -*- coding: utf-8 -*-
# """
# nirf_extract.py  -  NIRF score-card image -> Excel extractor

# Install:  pip install pytesseract openpyxl Pillow
# Tesseract binary: https://github.com/UB-Mannheim/tesseract/wiki

# How it works:
#   JPG (2025) -> thick dark borders -> pad=3, scale=3, standard grayscale OCR
#   PNG (2023) -> thin light borders -> pad=0, scale=8
#              -> per-column OCR using max-channel binarization (thresh=155)
#                 removes colored bar-chart bleed from score/total cells
#              -> smart_fix_decimal recovers missing decimal points

# Usage:
#   python scripts/paddle_extract.py <image_or_folder> [--output out.xlsx] [--workers N]
# """

# import io, os, sys, re, json, argparse
# from pathlib import Path
# from concurrent.futures import ThreadPoolExecutor, as_completed
# import multiprocessing

# if hasattr(sys.stdout, "reconfigure"):
#     sys.stdout.reconfigure(encoding="utf-8", errors="replace")
# else:
#     sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# # TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
# TESSERACT_CMD = None


# def setup_tesseract():
#     import pytesseract
#     if TESSERACT_CMD:
#         pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
#     else:
#         default = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
#         if os.path.exists(default):
#             pytesseract.pytesseract.tesseract_cmd = default
#     return pytesseract


# def open_image(path):
#     from PIL import Image
#     return Image.open(path).convert("RGB")

# def to_numpy(img):
#     import numpy as np
#     return np.array(img)


# # ── Border detection ──────────────────────────────────────────────────────────

# def find_dark_borders(arr):
#     """Thick dark horizontal lines — JPG/2025 style."""
#     import numpy as np
#     h, w = arr.shape[:2]
#     borders, in_b = [], False
#     for y in range(int(h * 0.25), h):
#         dark = int(np.sum((arr[y,:,0]<100)&(arr[y,:,1]<100)&(arr[y,:,2]<100)))
#         if dark > w * 0.3:
#             if not in_b: borders.append(y); in_b = True
#         else: in_b = False
#     return borders


# def find_light_borders(arr):
#     """Thin uniform light-gray lines — PNG/2023 style. Returns deduped list."""
#     import numpy as np
#     h, w = arr.shape[:2]
#     raw, in_b = [], False
#     for y in range(int(h * 0.25), h):
#         row = arr[y]
#         std = float(row.std(axis=0).mean())
#         nw  = int(np.sum((row[:,0]>200)&(row[:,1]>200)&(row[:,2]>200)))
#         if std < 15 and nw > w * 0.85:
#             if not in_b: raw.append(y); in_b = True
#         else: in_b = False
#     deduped = []
#     for b in raw:
#         if not deduped or b - deduped[-1] >= 20:
#             deduped.append(b)
#     return deduped


# def get_band_borders(arr):
#     """
#     Returns (img_type, pad, scale, b0, b1, b2, b3, b4).
#     JPG: pad=3, scale=3
#     PNG: pad=0, scale=8
#     """
#     import numpy as np
#     h = arr.shape[0]

#     dark = find_dark_borders(arr)
#     if len(dark) >= 4:
#         b0, b1, b2, b3 = dark[0], dark[1], dark[2], dark[3]
#         b4 = min(b3 + 50, h)
#         return 'jpg', 3, 3, b0, b1, b2, b3, b4
#     else:
#         light = find_light_borders(arr)
#         inner = light[1:] if len(light) >= 6 else light
#         if len(inner) >= 5:
#             b0, b1, b2, b3, b4 = inner[0], inner[1], inner[2], inner[3], inner[4]
#         else:
#             b0 = int(h*0.323); b1 = int(h*0.357)
#             b2 = int(h*0.390); b3 = int(h*0.425); b4 = int(h*0.459)
#         return 'png', 0, 8, b0, b1, b2, b3, b4


# # ── Column header OCR (full-band) ─────────────────────────────────────────────

# def ocr_headers(pytesseract, pil_crop, scale):
#     """OCR header band. Returns [(x_center, text), ...] sorted left to right."""
#     from PIL import Image
#     import numpy as np

#     w, h = pil_crop.size
#     if w < 1 or h < 1:
#         return []

#     up   = pil_crop.resize((w * scale, h * scale), Image.LANCZOS)
#     gray = up.convert("L")
#     arr  = np.array(gray)
#     if arr.mean() < 128:
#         gray = gray.point(lambda x: 255 - x)

#     config = "--psm 11 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ+.0123456789"
#     data   = pytesseract.image_to_data(gray, config=config,
#                                        output_type=pytesseract.Output.DICT)
#     items = []
#     for i, text in enumerate(data["text"]):
#         text = text.strip()
#         if not text or data["conf"][i] < 10:
#             continue
#         xc = (data["left"][i] + data["width"][i] / 2) / scale
#         items.append((xc, text))

#     items.sort(key=lambda t: t[0])
#     return items


# # ── Per-column value OCR (max-channel binarization) ───────────────────────────

# def ocr_col_value(arr, t_band, b_band, x1, x2, img_w, scale=8, thresh=155):
#     """
#     OCR a single score/total cell using max-channel binarization.

#     Max-channel binarization:  for each pixel, take max(R,G,B).
#     If max < thresh -> black (text pixel).  Else -> white (background).

#     This cleanly separates near-black text from colored bar-chart bleed
#     even when the bar overlaps the score cell.

#     thresh=155:  experimentally found to capture all digit strokes while
#                  eliminating colored backgrounds.
#     """
#     from PIL import Image
#     import numpy as np

#     margin = (x2 - x1) * 0.3     # 30% overlap into neighbors for context
#     x1w = max(0,   int(x1 - margin))
#     x2w = min(img_w, int(x2 + margin))

#     cell = arr[t_band:b_band, x1w:x2w, :]
#     if cell.shape[0] < 3 or cell.shape[1] < 3:
#         return ''

#     pil = Image.fromarray(cell)
#     cw, ch = pil.size

#     # LANCZOS upscale preserves sub-pixel detail
#     up_rgb = pil.resize((cw * scale, ch * scale), Image.LANCZOS)
#     up_arr = np.array(up_rgb)

#     # Max-channel binarize: isolate near-black text
#     r  = up_arr[:, :, 0].astype(int)
#     g  = up_arr[:, :, 1].astype(int)
#     b  = up_arr[:, :, 2].astype(int)
#     max_ch = np.maximum(np.maximum(r, g), b)
#     clean  = np.where(max_ch < thresh, 0, 255).astype('uint8')

#     pil_clean = Image.fromarray(clean)
#     pytesseract = setup_tesseract()

#     data = pytesseract.image_to_data(
#         pil_clean,
#         config="--psm 11 -c tessedit_char_whitelist=.0123456789",
#         output_type=pytesseract.Output.DICT
#     )

#     # Collect all tokens, rank by proximity to column center
#     col_cx = (x1 + x2) / 2
#     items  = []
#     for i, text in enumerate(data["text"]):
#         text = text.strip()
#         if not text:
#             continue
#         xc = (data["left"][i] + data["width"][i] / 2) / scale + x1w
#         items.append((abs(xc - col_cx), text))

#     if not items:
#         return ''
#     items.sort()
#     return items[0][1]


# # ── JPG value OCR (standard grayscale full-band) ──────────────────────────────

# def ocr_band_jpg(pytesseract, pil_crop, scale):
#     """Standard grayscale PSM-11 OCR for JPG score/total bands."""
#     from PIL import Image
#     import numpy as np

#     w, h = pil_crop.size
#     if w < 1 or h < 1:
#         return []

#     up   = pil_crop.resize((w * scale, h * scale), Image.LANCZOS)
#     gray = up.convert("L")
#     arr  = np.array(gray)
#     if arr.mean() < 128:
#         gray = gray.point(lambda x: 255 - x)

#     config = "--psm 11 -c tessedit_char_whitelist=.0123456789"
#     data   = pytesseract.image_to_data(gray, config=config,
#                                        output_type=pytesseract.Output.DICT)
#     items = []
#     for i, text in enumerate(data["text"]):
#         text = text.strip()
#         if not text or data["conf"][i] < 10:
#             continue
#         xc = (data["left"][i] + data["width"][i] / 2) / scale
#         items.append((xc, text))

#     items.sort(key=lambda t: t[0])
#     return items


# # ── Post-processing ───────────────────────────────────────────────────────────

# def smart_fix_decimal(s):
#     """
#     Recover missing decimal point in NIRF score/total values.
#     NIRF format: always exactly 2 decimal places (X.XX or XX.XX).

#     '2747'  -> '27.47'  (4 digits: dd.dd)
#     '1760'  -> '17.60'  (4 digits: dd.dd)
#     '593'   -> '5.93'   (3 digits: d.dd)
#     '17.60' -> '17.60'  (unchanged — already correct)
#     """
#     s = s.strip()
#     if not s:
#         return s
#     if re.fullmatch(r'\d+\.\d+', s):
#         return s      # already has decimal
#     if s.isdigit():
#         n = len(s)
#         if n >= 3:
#             return s[:-2] + '.' + s[-2:]   # insert before last 2 digits
#         if n == 2:
#             return s[0] + '.' + s[1]
#     return s


# # ── Column header classification & OCR correction ────────────────────────────

# GROUP_HEADERS = {"TLR", "RP", "GO", "OI"}
# ROW_LABELS    = {"SCORE", "TOTAL", "RANK"}

# KNOWN_COLS = {
#     "SS", "FSR", "FQE", "FRU", "OE+MIR", "OE",
#     "PU", "QP", "IPR", "FPPP", "SDG",
#     "GPH", "GUE", "MS", "GPHD",
#     "RD", "WD", "ESCS", "PCS",
#     "PR", "PERCEPTION",
# }

# OCR_FIXES = {
#     "S5":  "SS",       "5S":  "SS",      "S$": "SS",     "55": "SS",
#     "W0":  "WD",       "VVD": "WD",      "WO": "WD",
#     "FQF": "FQE",      "FOE": "FQE",     "FDE": "FQE",
#     "FRO": "FRU",      "FR0": "FRU",
#     "1PR": "IPR",      "lPR": "IPR",
#     "GU3": "GUE",
#     "P5":  "PCS",      "PC5": "PCS",
#     "E5CS":"ESCS",     "ESC5":"ESCS",
#     "0E+MIR":"OE+MIR", "OE-MIR":"OE+MIR","OE+MR":"OE+MIR",
#     "5DG": "SDG",      "SDC": "SDG",
#     "M5":  "MS",       "GP H":"GPH",
#     "OP":  "QP",       "OF":  "OE",
# }

# def fix_ocr(tok):
#     t = tok.upper().replace(" ", "")
#     return OCR_FIXES.get(t, t)

# def is_float(s):
#     try: float(s); return True
#     except ValueError: return False

# def is_col_header(tok):
#     t = fix_ocr(tok)
#     if t in GROUP_HEADERS or t in ROW_LABELS:
#         return False
#     if t in KNOWN_COLS:
#         return True
#     return bool(re.fullmatch(r"[A-Z+]{2,8}", t))


# # ── Per-image extraction ──────────────────────────────────────────────────────

# def extract_one(image_path):
#     pytesseract = setup_tesseract()
#     import numpy as np
#     from PIL import Image as PILImage

#     img = open_image(image_path)
#     arr = to_numpy(img)
#     h, w = arr.shape[:2]

#     # ── Institute code ────────────────────────────────────────────────────────
#     fname = Path(image_path).stem
#     institute_code = fname if re.match(r"IR-[A-Z]-[A-Z]-\d+", fname) else ""

#     # ── Institute name ────────────────────────────────────────────────────────
#     title_start = title_end = None
#     for y in range(min(80, h)):
#         nw = int(np.sum(~((arr[y,:,0]>220)&(arr[y,:,1]>220)&(arr[y,:,2]>220))))
#         if nw > 20:
#             if title_start is None: title_start = y
#             title_end = y
#     if title_start is None:
#         title_start, title_end = 15, 32

#     t1, t2 = max(0, title_start - 4), min(h, title_end + 4)
#     title_crop = img.crop((0, t1, w, t2))
#     tw, th = title_crop.size
#     title_big  = title_crop.resize((tw * 4, th * 4), PILImage.LANCZOS)
#     title_text = pytesseract.image_to_string(title_big.convert("L"), config="--psm 7").strip()

#     m = re.search(r"(.+?)\s*\(IR-[A-Z]-[A-Z]-\d+\)", title_text)
#     if m:
#         institute_name = m.group(1).strip()
#     else:
#         lines = [l.strip() for l in title_text.splitlines() if len(l.strip()) > 4]
#         institute_name = max(lines, key=len) if lines else ""
#     institute_name = re.sub(r"^[^A-Za-z]+", "", institute_name).strip()

#     # ── Band borders ──────────────────────────────────────────────────────────
#     img_type, pad, scale, b0, b1, b2, b3, b4 = get_band_borders(arr)

#     # ── Crop bands ────────────────────────────────────────────────────────────
#     col_pil = img.crop((0, b1 + pad, w, b2 - pad))

#     # ── OCR column headers (full-band LANCZOS) ────────────────────────────────
#     col_items    = ocr_headers(pytesseract, col_pil, scale)
#     header_items = [(x, fix_ocr(t)) for x, t in col_items if is_col_header(t)]

#     if not header_items:
#         n_cols = n_scores = 0
#         table  = {}
#         return {
#             "image_path": image_path, "year": "", "category": "",
#             "institute_name": institute_name, "institute_code": institute_code,
#             "table": table,
#             "_msg": f"  [{img_type.upper()}] {institute_code} | {institute_name[:38] or '?'} | cols=0 scores=0 MISMATCH",
#         }

#     # ── Compute column x-boundaries ───────────────────────────────────────────
#     x_centers = [xc for xc, _ in header_items]
#     x_bounds  = (
#         [0]
#         + [(x_centers[i] + x_centers[i+1]) / 2 for i in range(len(x_centers) - 1)]
#         + [w]
#     )

#     # ── OCR score and total rows ───────────────────────────────────────────────
#     table = {}

#     if img_type == 'png':
#         # PNG: per-column max-channel binarization
#         # Removes colored bar-chart bleed that overlaps into score cells
#         for i, (hx, hname) in enumerate(header_items):
#             x1, x2 = x_bounds[i], x_bounds[i + 1]

#             score_raw = ocr_col_value(arr, b2, b3, x1, x2, w, scale=8, thresh=155)
#             total_raw = ocr_col_value(arr, b3, b4, x1, x2, w, scale=8, thresh=155)

#             score_val = smart_fix_decimal(score_raw)
#             total_val = smart_fix_decimal(total_raw)

#             table[hname] = {
#                 "score": score_val if is_float(score_val) else "",
#                 "total": total_val if is_float(total_val) else "",
#             }

#     else:
#         # JPG: full-band standard grayscale OCR
#         score_pil  = img.crop((0, b2 + pad, w, b3 - pad))
#         total_pil  = img.crop((0, b3 + pad, w, b4 - pad))

#         score_items = ocr_band_jpg(pytesseract, score_pil, scale)
#         total_items = ocr_band_jpg(pytesseract, total_pil, scale)

#         def match_by_x(header_items, value_items):
#             result, used = {}, set()
#             for hx, htext in header_items:
#                 best_dist, best_idx = float("inf"), None
#                 for i, (vx, _) in enumerate(value_items):
#                     if i in used: continue
#                     dist = abs(hx - vx)
#                     if dist < best_dist:
#                         best_dist, best_idx = dist, i
#                 if best_idx is not None:
#                     used.add(best_idx)
#                     result[htext] = value_items[best_idx][1]
#             return result

#         score_map = match_by_x(header_items, [(x, v) for x, v in score_items if is_float(v)])
#         total_map = match_by_x(header_items, [(x, v) for x, v in total_items if is_float(v)])

#         for _, hname in header_items:
#             table[hname] = {
#                 "score": score_map.get(hname, ""),
#                 "total": total_map.get(hname, ""),
#             }

#     # ── Year / category from folder structure ─────────────────────────────────
#     parts = Path(image_path).resolve().parts
#     year = category = ""
#     for i, p in enumerate(parts):
#         if p == "image" and i >= 2:
#             year, category = parts[i-2], parts[i-1]
#             break

#     n_cols   = len(table)
#     n_scores = sum(1 for v in table.values() if v["score"])
#     flag     = " MISMATCH" if n_cols != n_scores and n_cols > 0 else " OK"

#     return {
#         "image_path":     image_path,
#         "year":           year,
#         "category":       category,
#         "institute_name": institute_name,
#         "institute_code": institute_code,
#         "table":          table,
#         "_msg":           f"  [{img_type.upper()}] {institute_code} | {institute_name[:38] if institute_name else '?'} | cols={n_cols} scores={n_scores}{flag}",
#     }


# # ── Collect images ────────────────────────────────────────────────────────────

# IMAGE_EXTS = {".jpg", ".jpeg", ".png"}

# def collect_images(path):
#     p = Path(os.path.normpath(path))
#     if p.is_file() and p.suffix.lower() in IMAGE_EXTS:
#         return [str(p)]
#     if p.is_dir():
#         imgs = []
#         for ext in IMAGE_EXTS:
#             imgs.extend(str(f) for f in p.rglob(f"*{ext}"))
#         imgs.sort()
#         return imgs
#     print(f"Path not found: {p}", file=sys.stderr)
#     return []


# # ── Build Excel ───────────────────────────────────────────────────────────────

# def build_excel(rows, output_path):
#     from openpyxl import Workbook
#     from openpyxl.styles import Font, PatternFill, Alignment

#     header_fill = PatternFill("solid", fgColor="1F4E79")
#     header_font = Font(color="FFFFFF", bold=True)
#     alt_fill    = PatternFill("solid", fgColor="D9E1F2")

#     all_headers, seen_h = [], set()
#     for row in rows:
#         for hk in row.get("table", {}):
#             if hk not in seen_h:
#                 seen_h.add(hk); all_headers.append(hk)

#     fixed_cols = ["Year", "Category", "Institute Name", "Institute Code"]
#     all_cols   = (fixed_cols
#                   + [f"{h} Score" for h in all_headers]
#                   + [f"{h} Total" for h in all_headers])

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "NIRF Scores"

#     for ci, col in enumerate(all_cols, 1):
#         cell = ws.cell(row=1, column=ci, value=col)
#         cell.fill = header_fill
#         cell.font = header_font
#         cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#     ws.row_dimensions[1].height = 30

#     for ri, row in enumerate(rows, 2):
#         tbl  = row.get("table", {})
#         fill = alt_fill if ri % 2 == 0 else PatternFill()
#         vals = ([row.get("year",""), row.get("category",""),
#                  row.get("institute_name",""), row.get("institute_code","")]
#                 + [tbl.get(h, {}).get("score", "") for h in all_headers]
#                 + [tbl.get(h, {}).get("total", "") for h in all_headers])
#         for ci, val in enumerate(vals, 1):
#             cell = ws.cell(row=ri, column=ci, value=val)
#             cell.fill = fill
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#     for ci in range(1, len(all_cols) + 1):
#         col_letter = ws.cell(row=1, column=ci).column_letter
#         max_len = max(
#             len(str(ws.cell(row=1, column=ci).value or "")),
#             *(len(str(ws.cell(row=ri, column=ci).value or "")) for ri in range(2, len(rows)+2))
#         )
#         ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

#     ws.freeze_panes = "A2"
#     wb.save(output_path)
#     print(f"\nDone. Saved {len(rows)} rows -> {output_path}", flush=True)


# # ── Main ──────────────────────────────────────────────────────────────────────

# def main():
#     parser = argparse.ArgumentParser(description="Extract NIRF score-card table data")
#     parser.add_argument("input",     help="Image file or folder (recursive)")
#     parser.add_argument("--output",  "-o", default="nirf_scores.xlsx")
#     parser.add_argument("--json",    "-j", default=None)
#     parser.add_argument("--workers", "-w", type=int,
#                         default=min(8, multiprocessing.cpu_count()),
#                         help="Parallel threads (default: min(8, cpu_count))")
#     args = parser.parse_args()

#     images = collect_images(args.input)
#     if not images:
#         print(f"No images found at: {args.input}", file=sys.stderr)
#         sys.exit(1)

#     n = len(images)
#     w = min(args.workers, n)
#     print(f"Found {n} image(s). Running with {w} thread(s)...\n", flush=True)

#     try:
#         pyt = setup_tesseract()
#         from PIL import Image as PILImage
#         import numpy as np
#         pyt.image_to_string(PILImage.fromarray(np.ones((10,10,3), dtype=np.uint8)*255))
#         print("Tesseract ready.\n", flush=True)
#     except Exception as e:
#         print(f"\nERROR: Tesseract not found: {e}", file=sys.stderr)
#         print("Install from: https://github.com/UB-Mannheim/tesseract/wiki", file=sys.stderr)
#         sys.exit(1)

#     rows_by_path = {}
#     done = 0

#     with ThreadPoolExecutor(max_workers=w) as exe:
#         futures = {exe.submit(extract_one, p): p for p in images}
#         for fut in as_completed(futures):
#             done += 1
#             img_path = futures[fut]
#             try:
#                 row = fut.result()
#                 msg = row.pop("_msg", "")
#                 rows_by_path[img_path] = row
#                 print(f"[{done}/{n}] {Path(img_path).name}{msg}", flush=True)
#             except Exception as e:
#                 print(f"[{done}/{n}] ERROR {Path(img_path).name}: {e}", flush=True)

#     rows = [rows_by_path[p] for p in images if p in rows_by_path]

#     if args.json:
#         with open(args.json, "w", encoding="utf-8") as f:
#             json.dump(rows, f, indent=2, ensure_ascii=False)
#         print(f"JSON -> {args.json}", flush=True)

#     if not rows:
#         print("No rows extracted.", flush=True)
#         sys.exit(0)

#     build_excel(rows, args.output)


# if __name__ == "__main__":
#     main()
















# working completly for 2023



# # -*- coding: utf-8 -*-
# """
# paddle_extract.py  —  NIRF score-card image → Excel extractor (v4)

# Improvements over previous versions
# ────────────────────────────────────
# 1. ROBUST LAYOUT DETECTION
#    • Locates the single thick dark separator line that divides the chart
#      from the data table. Everything else is anchored relative to it.
#    • No more hard-coded y-fractions that break on different image heights.

# 2. THREE-PASS COLUMN HEADER READING
#    Pass A – Full-band PSM 11 on the header strip (catches most columns).
#    Pass B – Per-column PSM 6 crop for any column that was missed or
#             truncated in Pass A (handles colored-text columns like QP, RD).
#    Pass C – Fuzzy OCR-error correction via a known-column dictionary +
#             Levenshtein-style character fixes.

# 3. SCORE/TOTAL ROW LOCATION via LABEL SEARCH
#    • Searches for the "Score" text label in the left margin to pinpoint
#      the exact y-centre of the score row (±12 px), then offsets for Total.
#    • Falls back to a separator-relative offset when the label is absent.

# 4. X-POSITION MATCHING
#    • Maps header x-positions → value x-positions using nearest-neighbour,
#      so a missing header OCR word cannot shift all downstream columns.

# 5. SMART DECIMAL FIX
#    • Recovers "1391" → "13.91" for PNG images where the decimal dot is
#      swallowed by bar-chart bleed.

# 6. INSTITUTE NAME / CODE EXTRACTION
#    • Code comes from filename (always reliable).
#    • Name is parsed from the image title row using an upscaled crop.

# Usage:
#     pip install pytesseract openpyxl Pillow numpy
#     Tesseract: https://github.com/UB-Mannheim/tesseract/wiki

#     python paddle_extract.py <image_or_folder> [--output out.xlsx] [--json out.json] [--workers N]
# """

# import io, os, sys, re, json, argparse
# from pathlib import Path
# from concurrent.futures import ThreadPoolExecutor, as_completed
# import multiprocessing

# if hasattr(sys.stdout, "reconfigure"):
#     sys.stdout.reconfigure(encoding="utf-8", errors="replace")
# else:
#     sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# TESSERACT_CMD = None   # set to r"C:\Program Files\Tesseract-OCR\tesseract.exe" if needed

# # ─────────────────────────────────────────────────────────────────────────────
# # Tesseract setup
# # ─────────────────────────────────────────────────────────────────────────────

# def setup_tesseract():
#     import pytesseract
#     if TESSERACT_CMD:
#         pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
#     else:
#         default = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
#         if os.path.exists(default):
#             pytesseract.pytesseract.tesseract_cmd = default
#     return pytesseract


# # ─────────────────────────────────────────────────────────────────────────────
# # Known NIRF column metadata
# # ─────────────────────────────────────────────────────────────────────────────

# # All known valid column abbreviations across all NIRF categories
# KNOWN_COLS = {
#     "SS", "FSR", "FQE", "FRU",
#     "OE+MIR", "OE", "MIR",
#     "PU", "QP", "IPR", "FPPP", "SDG",
#     "GPH", "GUE", "MS", "GPHD",
#     "RD", "WD", "ESCS", "PCS",
#     "PR", "PERCEPTION",
# }

# GROUP_HEADERS = {"TLR", "RP", "GO", "OI"}   # NOTE: PR is NOT here; it is both a group label and a data column
# ROW_LABELS    = {"SCORE", "TOTAL", "RANK"}

# # Common Tesseract misreads → correct token
# OCR_FIXES = {
#     "S5":       "SS",      "5S":      "SS",      "S$":    "SS",    "55":   "SS",
#     "W0":       "WD",      "VVD":     "WD",       "WO":   "WD",
#     "FQF":      "FQE",     "FOE":     "FQE",      "FDE":  "FQE",
#     "FRO":      "FRU",     "FR0":     "FRU",
#     "1PR":      "IPR",     "lPR":     "IPR",
#     "GU3":      "GUE",
#     "P5":       "PCS",     "PC5":     "PCS",
#     "E5CS":     "ESCS",    "ESC5":    "ESCS",     "ECSS": "ESCS",
#     "0E+MIR":   "OE+MIR",  "OE-MIR":  "OE+MIR",  "OE+MR":"OE+MIR",
#     "5DG":      "SDG",     "SDC":     "SDG",
#     "M5":       "MS",      "GP H":    "GPH",
#     "OP":       "QP",      "OF":      "OE",
#     "GPHD":     "GPHD",    "GPHO":    "GPHD",
#     "SR":       "FSR",     "FSP":     "FSR",
#     "OE":       "OE",      "MlR":     "MIR",
#     "FRl":      "FRU",     "FRI":     "FRU",
# }

# # Prefix-completion: when only the last part of an abbreviation is captured
# PREFIX_COMPLETIONS = {
#     "SR":   "FSR",
#     "QE":   "FQE",
#     "RU":   "FRU",
#     "S":    "SS",    # lone S → SS (only when other cols confirm SS is missing)
#     "PH":   "GPH",
#     "UE":   "GUE",
#     "HD":   "GPHD",
#     "CS":   "PCS",   # careful: also could be ESCS
#     "SCS":  "ESCS",
# }


# def fix_ocr_token(tok: str) -> str:
#     """Apply OCR error fixes and return the corrected token."""
#     t = tok.upper().replace(" ", "")
#     if t in OCR_FIXES:
#         return OCR_FIXES[t]
#     return t


# def is_col_header(tok: str) -> bool:
#     t = fix_ocr_token(tok)
#     if t in GROUP_HEADERS or t in ROW_LABELS:
#         return False
#     if t in KNOWN_COLS:
#         return True
#     # Accept any 2–8 uppercase letter sequence not obviously noise
#     return bool(re.fullmatch(r"[A-Z+]{2,8}", t))


# # ─────────────────────────────────────────────────────────────────────────────
# # Layout detection
# # ─────────────────────────────────────────────────────────────────────────────

# def find_dark_separator(arr: "np.ndarray", start_frac: float = 0.35) -> int | None:
#     """
#     Find the single thick black horizontal line that separates the bar chart
#     from the data table.  Returns the y-coordinate or None.
#     """
#     import numpy as np
#     h, w = arr.shape[:2]
#     for y in range(int(h * start_frac), h):
#         dark = int(np.sum((arr[y, :, 0] < 80) & (arr[y, :, 1] < 80) & (arr[y, :, 2] < 80)))
#         if dark > w * 0.75:
#             return y
#     return None


# def find_score_total_y(img, sep_y: int, w: int):
#     """
#     Search for the "Score" label in the left margin to pinpoint exact row y.
#     Returns (score_y, total_y) as floats.
#     """
#     pytesseract = setup_tesseract()
#     search_top  = sep_y + 40
#     search_bot  = sep_y + 200
#     # Only scan the left ~200px where the label lives
#     region = img.crop((0, search_top, min(200, w), search_bot))
#     scale  = 4
#     region_big = region.resize((region.size[0] * scale, region.size[1] * scale), _lanczos())
#     data = pytesseract.image_to_data(
#         region_big.convert("L"),
#         config="--psm 11",
#         output_type=pytesseract.Output.DICT,
#     )
#     score_y = total_y = None
#     for i, text in enumerate(data["text"]):
#         t = text.strip().lower()
#         if "score" in t and score_y is None:
#             score_y = (data["top"][i] + data["height"][i] / 2) / scale + search_top
#         if "total" in t and total_y is None:
#             total_y = (data["top"][i] + data["height"][i] / 2) / scale + search_top

#     if score_y is None:
#         score_y = sep_y + 80
#     if total_y is None:
#         total_y = score_y + 32

#     return score_y, total_y


# def _lanczos():
#     from PIL import Image
#     return Image.LANCZOS


# # ─────────────────────────────────────────────────────────────────────────────
# # Column header OCR — three-pass
# # ─────────────────────────────────────────────────────────────────────────────

# def ocr_full_header_band(pytesseract, img, sep_y: int, w: int, scale: int = 4):
#     """Pass A: full-band PSM 11 on the column header strip."""
#     from PIL import Image
#     crop = img.crop((0, sep_y - 38, w, sep_y))
#     big  = crop.resize((w * scale, crop.size[1] * scale), Image.LANCZOS)
#     gray = big.convert("L")
#     data = pytesseract.image_to_data(
#         gray,
#         config="--psm 11 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ+",
#         output_type=pytesseract.Output.DICT,
#     )
#     items = []
#     for i, text in enumerate(data["text"]):
#         text = text.strip()
#         if not text or data["conf"][i] < 5:
#             continue
#         xc = (data["left"][i] + data["width"][i] / 2) / scale
#         items.append((xc, text))
#     items.sort(key=lambda t: t[0])
#     return items


# def ocr_per_column_header(pytesseract, img, xc: float, sep_y: int, w: int, scale: int = 6):
#     """Pass B: narrow crop centred on one column, high upscale."""
#     from PIL import Image
#     margin = 48
#     x1 = max(0, int(xc - margin))
#     x2 = min(w, int(xc + margin))
#     crop = img.crop((x1, sep_y - 38, x2, sep_y))
#     big  = crop.resize((crop.size[0] * scale, crop.size[1] * scale), Image.LANCZOS)
#     gray = big.convert("L")
#     text = pytesseract.image_to_string(
#         gray,
#         config="--psm 6 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ+",
#     ).strip().replace("\n", " ").strip()
#     # Take the longest token
#     tokens = [t for t in text.split() if t]
#     return max(tokens, key=len) if tokens else ""


# def build_header_map(pytesseract, img, sep_y: int, w: int, score_xs: list[float]) -> dict[float, str]:
#     """
#     Return {score_x: col_name} using three-pass OCR + fuzzy correction.
#     score_xs are the x-centres of score values (reliable ground truth for column positions).
#     """

#     # ── Pass A ───────────────────────────────────────────────────────────────
#     raw_items = ocr_full_header_band(pytesseract, img, sep_y, w)
#     # Filter to plausible column headers
#     pass_a = [(x, fix_ocr_token(t)) for x, t in raw_items if is_col_header(t)]

#     # ── Map Pass A results to score_xs by nearest neighbour ──────────────────
#     assigned: dict[int, str] = {}  # score_xs index → col name
#     used_score = set()
#     for hx, hname in pass_a:
#         best_dist, best_idx = float("inf"), None
#         for i, sx in enumerate(score_xs):
#             if i in used_score:
#                 continue
#             d = abs(hx - sx)
#             if d < best_dist:
#                 best_dist, best_idx = d, i
#         if best_idx is not None and best_dist < 80:
#             assigned[best_idx] = hname
#             used_score.add(best_idx)

#     # ── Pass B: fill gaps ─────────────────────────────────────────────────────
#     missing = [i for i in range(len(score_xs)) if i not in assigned]
#     for i in missing:
#         xc = score_xs[i]
#         raw = ocr_per_column_header(pytesseract, img, xc, sep_y, w)
#         if raw:
#             fixed = fix_ocr_token(raw)
#             # Try prefix completion if still not in KNOWN_COLS
#             if fixed not in KNOWN_COLS and fixed in PREFIX_COMPLETIONS:
#                 fixed = PREFIX_COMPLETIONS[fixed]
#             if is_col_header(fixed):
#                 assigned[i] = fixed

#     # ── Pass C: prefix-complete any truncated tokens ──────────────────────────
#     for i, name in list(assigned.items()):
#         if name not in KNOWN_COLS and name in PREFIX_COMPLETIONS:
#             assigned[i] = PREFIX_COMPLETIONS[name]

#     # Build final dict keyed by score_x
#     return {score_xs[i]: assigned.get(i, f"COL{i+1}") for i in range(len(score_xs))}


# # ─────────────────────────────────────────────────────────────────────────────
# # Value extraction
# # ─────────────────────────────────────────────────────────────────────────────

# def extract_values_at_y(pytesseract, img, y_centre: float, w: int, scale: int = 4):
#     """
#     Extract (x_centre, value_string) pairs from a horizontal data row.
#     Returns floats only.
#     """
#     import numpy as np
#     from PIL import Image

#     half = 14
#     y1, y2 = max(0, int(y_centre - half)), int(y_centre + half)
#     crop = img.crop((80, y1, w, y2))    # skip left label area
#     big  = crop.resize((crop.size[0] * scale, crop.size[1] * scale), Image.LANCZOS)

#     arr  = np.array(big)
#     # Max-channel binarize to strip bar-chart bleed
#     max_ch = np.maximum(np.maximum(arr[:,:,0].astype(int), arr[:,:,1].astype(int)), arr[:,:,2].astype(int))
#     clean  = np.where(max_ch < 160, 0, 255).astype("uint8")
#     from PIL import Image as PILImage
#     pil_clean = PILImage.fromarray(clean)

#     data = pytesseract.image_to_data(
#         pil_clean,
#         config="--psm 11 -c tessedit_char_whitelist=.0123456789",
#         output_type=pytesseract.Output.DICT,
#     )
#     items = []
#     for i, text in enumerate(data["text"]):
#         text = text.strip()
#         if not text or data["conf"][i] < 5:
#             continue
#         xc = (data["left"][i] + data["width"][i] / 2) / scale + 80
#         items.append((xc, text))
#     items.sort(key=lambda t: t[0])
#     return items


# def smart_fix_decimal(s: str) -> str:
#     """
#     Recover a missing decimal point: '1391' → '13.91', '2765' → '27.65'.
#     NIRF values always have exactly 2 decimal places.
#     """
#     s = s.strip()
#     if not s:
#         return s
#     if re.fullmatch(r"\d+\.\d+", s):
#         return s   # already correct
#     if s.isdigit() and len(s) >= 3:
#         return s[:-2] + "." + s[-2:]
#     return s


# def is_float(s: str) -> bool:
#     try:
#         float(s)
#         return True
#     except ValueError:
#         return False


# def match_values_to_headers(
#     header_map: dict[float, str],
#     value_items: list[tuple[float, str]],
# ) -> dict[str, str]:
#     """Nearest-neighbour match between header x-positions and value x-positions."""
#     result:  dict[str, str] = {}
#     used = set()
#     for hx, col_name in sorted(header_map.items()):
#         best_dist, best_val = float("inf"), ""
#         for i, (vx, vtext) in enumerate(value_items):
#             if i in used:
#                 continue
#             d = abs(hx - vx)
#             if d < best_dist:
#                 best_dist, best_val = d, vtext
#                 best_idx = i
#         if best_dist < 100:
#             used.add(best_idx)
#             result[col_name] = smart_fix_decimal(best_val)
#     return result


# # ─────────────────────────────────────────────────────────────────────────────
# # Institute name / code
# # ─────────────────────────────────────────────────────────────────────────────

# def extract_institute_meta(pytesseract, img, arr: "np.ndarray"):
#     import numpy as np
#     from PIL import Image
#     h, w = arr.shape[:2]

#     # Code from filename (always reliable)
#     fname = Path(getattr(img, "filename", "")).stem
#     institute_code = fname if re.match(r"IR-[A-Z]-[A-Z]-\d+", fname) else ""

#     # Find title rows: scan top 90 px for non-white content
#     t_start = t_end = None
#     for y in range(min(90, h)):
#         nw = int(np.sum(~((arr[y,:,0] > 210) & (arr[y,:,1] > 210) & (arr[y,:,2] > 210))))
#         if nw > 30:
#             if t_start is None:
#                 t_start = y
#             t_end = y

#     if t_start is None:
#         t_start, t_end = 5, 30

#     title_crop = img.crop((0, max(0, t_start - 4), w, min(h, t_end + 6)))
#     tw, th = title_crop.size
#     title_big = title_crop.resize((tw * 5, th * 5), Image.LANCZOS)
#     title_text = pytesseract.image_to_string(title_big.convert("L"), config="--psm 7").strip()

#     m = re.search(r"(.+?)\s*\(IR-[A-Z]-[A-Z]-\d+\)", title_text)
#     if m:
#         institute_name = m.group(1).strip()
#     else:
#         lines = [l.strip() for l in title_text.splitlines() if len(l.strip()) > 4]
#         institute_name = max(lines, key=len) if lines else ""
#     institute_name = re.sub(r"^[^A-Za-z;]+", "", institute_name).strip().lstrip(";").strip()

#     return institute_name, institute_code


# # ─────────────────────────────────────────────────────────────────────────────
# # Per-image extraction entry point
# # ─────────────────────────────────────────────────────────────────────────────

# def extract_one(image_path: str) -> dict:
#     from PIL import Image
#     import numpy as np

#     pytesseract = setup_tesseract()

#     img = Image.open(image_path).convert("RGB")
#     img.filename = image_path   # keep for meta extraction
#     arr = np.array(img)
#     h, w = arr.shape[:2]

#     # ── Institute meta ────────────────────────────────────────────────────────
#     institute_name, institute_code = extract_institute_meta(pytesseract, img, arr)

#     # ── Find layout anchor ────────────────────────────────────────────────────
#     sep_y = find_dark_separator(arr)
#     if sep_y is None:
#         # Graceful degradation: no separator found
#         return {
#             "image_path": image_path, "year": "", "category": "",
#             "institute_name": institute_name, "institute_code": institute_code,
#             "table": {},
#             "_msg": f"  NO_SEP {institute_code} | {institute_name[:40]}",
#         }

#     # ── Find Score / Total row y-positions ───────────────────────────────────
#     score_y, total_y = find_score_total_y(img, sep_y, w)

#     # ── Extract score values (ground-truth x-positions) ───────────────────────
#     score_items = extract_values_at_y(pytesseract, img, score_y, w)
#     score_float = [(x, smart_fix_decimal(v)) for x, v in score_items if is_float(smart_fix_decimal(v))]
#     score_xs = [x for x, _ in score_float]

#     if not score_xs:
#         return {
#             "image_path": image_path, "year": "", "category": "",
#             "institute_name": institute_name, "institute_code": institute_code,
#             "table": {},
#             "_msg": f"  NO_SCORES {institute_code} | {institute_name[:40]}",
#         }

#     # ── Build column header map ───────────────────────────────────────────────
#     header_map = build_header_map(pytesseract, img, sep_y, w, score_xs)

#     # ── Extract total values ───────────────────────────────────────────────────
#     total_items = extract_values_at_y(pytesseract, img, total_y, w)
#     total_float = [(x, smart_fix_decimal(v)) for x, v in total_items if is_float(smart_fix_decimal(v))]

#     # ── Match values to headers ───────────────────────────────────────────────
#     score_map = match_values_to_headers(header_map, score_float)
#     total_map = match_values_to_headers(header_map, total_float)

#     table = {
#         col_name: {
#             "score": score_map.get(col_name, ""),
#             "total": total_map.get(col_name, ""),
#         }
#         for col_name in header_map.values()
#     }

#     # ── Year / category from folder structure ─────────────────────────────────
#     parts = Path(image_path).resolve().parts
#     year = category = ""
#     for i, p in enumerate(parts):
#         if p == "image" and i >= 2:
#             year, category = parts[i - 2], parts[i - 1]
#             break

#     n_cols   = len(table)
#     n_scores = sum(1 for v in table.values() if v["score"])
#     flag     = " MISMATCH" if n_cols > 0 and n_cols != n_scores else " OK"

#     return {
#         "image_path":     image_path,
#         "year":           year,
#         "category":       category,
#         "institute_name": institute_name,
#         "institute_code": institute_code,
#         "table":          table,
#         "_msg":           (
#             f"  {institute_code} | {(institute_name or '?')[:40]} "
#             f"| cols={n_cols} scores={n_scores}{flag}"
#         ),
#     }


# # ─────────────────────────────────────────────────────────────────────────────
# # Image collection
# # ─────────────────────────────────────────────────────────────────────────────

# IMAGE_EXTS = {".jpg", ".jpeg", ".png"}


# def collect_images(path: str) -> list[str]:
#     p = Path(os.path.normpath(path))
#     if p.is_file() and p.suffix.lower() in IMAGE_EXTS:
#         return [str(p)]
#     if p.is_dir():
#         imgs = []
#         for ext in IMAGE_EXTS:
#             imgs.extend(str(f) for f in p.rglob(f"*{ext}"))
#         imgs.sort()
#         return imgs
#     print(f"Path not found: {p}", file=sys.stderr)
#     return []


# # ─────────────────────────────────────────────────────────────────────────────
# # Excel output
# # ─────────────────────────────────────────────────────────────────────────────

# def build_excel(rows: list[dict], output_path: str) -> None:
#     from openpyxl import Workbook
#     from openpyxl.styles import Font, PatternFill, Alignment

#     header_fill  = PatternFill("solid", fgColor="1F4E79")
#     header_font  = Font(color="FFFFFF", bold=True)
#     alt_fill     = PatternFill("solid", fgColor="D9E1F2")
#     center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

#     all_headers: list[str] = []
#     seen_h: set[str] = set()
#     for row in rows:
#         for hk in row.get("table", {}):
#             if hk not in seen_h:
#                 seen_h.add(hk)
#                 all_headers.append(hk)

#     fixed_cols = ["Year", "Category", "Institute Name", "Institute Code"]
#     all_cols   = (
#         fixed_cols
#         + [f"{h} Score" for h in all_headers]
#         + [f"{h} Total" for h in all_headers]
#     )

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "NIRF Scores"

#     for ci, col in enumerate(all_cols, 1):
#         cell = ws.cell(row=1, column=ci, value=col)
#         cell.fill  = header_fill
#         cell.font  = header_font
#         cell.alignment = center_align
#     ws.row_dimensions[1].height = 30

#     for ri, row in enumerate(rows, 2):
#         tbl  = row.get("table", {})
#         fill = alt_fill if ri % 2 == 0 else PatternFill()
#         vals = (
#             [row.get("year",""), row.get("category",""),
#              row.get("institute_name",""), row.get("institute_code","")]
#             + [tbl.get(h, {}).get("score", "") for h in all_headers]
#             + [tbl.get(h, {}).get("total", "") for h in all_headers]
#         )
#         for ci, val in enumerate(vals, 1):
#             cell = ws.cell(row=ri, column=ci, value=val)
#             cell.fill      = fill
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#     for ci in range(1, len(all_cols) + 1):
#         col_letter = ws.cell(row=1, column=ci).column_letter
#         max_len = max(
#             len(str(ws.cell(row=1, column=ci).value or "")),
#             *(len(str(ws.cell(row=ri, column=ci).value or "")) for ri in range(2, len(rows) + 2)),
#         )
#         ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

#     ws.freeze_panes = "A2"
#     wb.save(output_path)
#     print(f"\nDone. Saved {len(rows)} rows → {output_path}", flush=True)


# # ─────────────────────────────────────────────────────────────────────────────
# # Main
# # ─────────────────────────────────────────────────────────────────────────────

# def main() -> None:
#     parser = argparse.ArgumentParser(description="Extract NIRF score-card table data (v4)")
#     parser.add_argument("input",     help="Image file or folder (searched recursively)")
#     parser.add_argument("--output",  "-o", default="nirf_scores.xlsx", help="Output Excel file")
#     parser.add_argument("--json",    "-j", default=None,               help="Also write JSON output")
#     parser.add_argument(
#         "--workers", "-w",
#         type=int,
#         default=min(8, multiprocessing.cpu_count()),
#         help="Parallel threads (default: min(8, cpu_count))",
#     )
#     args = parser.parse_args()

#     images = collect_images(args.input)
#     if not images:
#         print(f"No images found at: {args.input}", file=sys.stderr)
#         sys.exit(1)

#     n = len(images)
#     w = min(args.workers, n)
#     print(f"Found {n} image(s).  Threads: {w}\n", flush=True)

#     # Quick sanity-check for Tesseract
#     try:
#         pyt = setup_tesseract()
#         from PIL import Image as PILImage
#         import numpy as np
#         pyt.image_to_string(PILImage.fromarray(np.ones((10, 10, 3), dtype=np.uint8) * 255))
#         print("Tesseract OK.\n", flush=True)
#     except Exception as e:
#         print(f"\nERROR: Tesseract not available: {e}", file=sys.stderr)
#         print("Install from: https://github.com/UB-Mannheim/tesseract/wiki", file=sys.stderr)
#         sys.exit(1)

#     rows_by_path: dict[str, dict] = {}
#     done = 0

#     with ThreadPoolExecutor(max_workers=w) as exe:
#         futures = {exe.submit(extract_one, p): p for p in images}
#         for fut in as_completed(futures):
#             done += 1
#             img_path = futures[fut]
#             try:
#                 row = fut.result()
#                 msg = row.pop("_msg", "")
#                 rows_by_path[img_path] = row
#                 print(f"[{done}/{n}] {Path(img_path).name}{msg}", flush=True)
#             except Exception as e:
#                 print(f"[{done}/{n}] ERROR {Path(img_path).name}: {e}", flush=True)

#     rows = [rows_by_path[p] for p in images if p in rows_by_path]

#     if args.json:
#         with open(args.json, "w", encoding="utf-8") as f:
#             json.dump(rows, f, indent=2, ensure_ascii=False)
#         print(f"JSON → {args.json}", flush=True)

#     if not rows:
#         print("No rows extracted.", flush=True)
#         sys.exit(0)

#     build_excel(rows, args.output)


# if __name__ == "__main__":
#     main()

























#PERFECT CODE FOR EVERY YEAR









# # -*- coding: utf-8 -*-
# """
# paddle_extract.py  —  NIRF score-card image → Excel extractor (v4)

# Improvements over previous versions
# ────────────────────────────────────
# 1. ROBUST LAYOUT DETECTION
#    • Locates the single thick dark separator line that divides the chart
#      from the data table. Everything else is anchored relative to it.
#    • No more hard-coded y-fractions that break on different image heights.

# 2. THREE-PASS COLUMN HEADER READING
#    Pass A – Full-band PSM 11 on the header strip (catches most columns).
#    Pass B – Per-column PSM 6 crop for any column that was missed or
#             truncated in Pass A (handles colored-text columns like QP, RD).
#    Pass C – Fuzzy OCR-error correction via a known-column dictionary +
#             Levenshtein-style character fixes.

# 3. SCORE/TOTAL ROW LOCATION via LABEL SEARCH
#    • Searches for the "Score" text label in the left margin to pinpoint
#      the exact y-centre of the score row (±12 px), then offsets for Total.
#    • Falls back to a separator-relative offset when the label is absent.

# 4. X-POSITION MATCHING
#    • Maps header x-positions → value x-positions using nearest-neighbour,
#      so a missing header OCR word cannot shift all downstream columns.

# 5. SMART DECIMAL FIX
#    • Recovers "1391" → "13.91" for PNG images where the decimal dot is
#      swallowed by bar-chart bleed.

# 6. INSTITUTE NAME / CODE EXTRACTION
#    • Code comes from filename (always reliable).
#    • Name is parsed from the image title row using an upscaled crop.

# Usage:
#     pip install pytesseract openpyxl Pillow numpy
#     Tesseract: https://github.com/UB-Mannheim/tesseract/wiki

#     python paddle_extract.py <image_or_folder> [--output out.xlsx] [--json out.json] [--workers N]
# """

# import io, os, sys, re, json, argparse
# from pathlib import Path
# from concurrent.futures import ThreadPoolExecutor, as_completed
# import multiprocessing

# if hasattr(sys.stdout, "reconfigure"):
#     sys.stdout.reconfigure(encoding="utf-8", errors="replace")
# else:
#     sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# TESSERACT_CMD = None   # set to r"C:\Program Files\Tesseract-OCR\tesseract.exe" if needed

# # ─────────────────────────────────────────────────────────────────────────────
# # Tesseract setup
# # ─────────────────────────────────────────────────────────────────────────────

# def setup_tesseract():
#     import pytesseract
#     if TESSERACT_CMD:
#         pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
#     else:
#         default = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
#         if os.path.exists(default):
#             pytesseract.pytesseract.tesseract_cmd = default
#     return pytesseract


# # ─────────────────────────────────────────────────────────────────────────────
# # Known NIRF column metadata
# # ─────────────────────────────────────────────────────────────────────────────

# # All known valid column abbreviations across all NIRF categories
# KNOWN_COLS = {
#     "SS", "FSR", "FQE", "FRU",
#     "OE+MIR", "OE", "MIR",
#     "PU", "QP", "IPR", "FPPP", "SDG",
#     "GPH", "GUE", "MS", "GPHD",
#     "RD", "WD", "ESCS", "PCS",
#     "PR", "PERCEPTION",
# }

# GROUP_HEADERS = {"TLR", "RP", "GO", "OI"}   # NOTE: PR is NOT here; it is both a group label and a data column
# ROW_LABELS    = {"SCORE", "TOTAL", "RANK"}

# # Common Tesseract misreads → correct token
# OCR_FIXES = {
#     "S5":       "SS",      "5S":      "SS",      "S$":    "SS",    "55":   "SS",
#     "W0":       "WD",      "VVD":     "WD",       "WO":   "WD",
#     "FQF":      "FQE",     "FOE":     "FQE",      "FDE":  "FQE",
#     "FRO":      "FRU",     "FR0":     "FRU",
#     "1PR":      "IPR",     "lPR":     "IPR",
#     "GU3":      "GUE",
#     "P5":       "PCS",     "PC5":     "PCS",
#     "E5CS":     "ESCS",    "ESC5":    "ESCS",     "ECSS": "ESCS",
#     "0E+MIR":   "OE+MIR",  "OE-MIR":  "OE+MIR",  "OE+MR":"OE+MIR",
#     "5DG":      "SDG",     "SDC":     "SDG",
#     "M5":       "MS",      "GP H":    "GPH",
#     "OP":       "QP",      "OF":      "OE",
#     "GPHD":     "GPHD",    "GPHO":    "GPHD",
#     "SR":       "FSR",     "FSP":     "FSR",
#     "OE":       "OE",      "MlR":     "MIR",
#     "FRl":      "FRU",     "FRI":     "FRU",
# }

# # Prefix-completion: when only the last part of an abbreviation is captured
# PREFIX_COMPLETIONS = {
#     "SR":   "FSR",
#     "QE":   "FQE",
#     "RU":   "FRU",
#     "S":    "SS",    # lone S → SS (only when other cols confirm SS is missing)
#     "PH":   "GPH",
#     "UE":   "GUE",
#     "HD":   "GPHD",
#     "CS":   "PCS",   # careful: also could be ESCS
#     "SCS":  "ESCS",
# }


# def fix_ocr_token(tok: str) -> str:
#     """Apply OCR error fixes and return the corrected token."""
#     t = tok.upper().replace(" ", "")
#     if t in OCR_FIXES:
#         return OCR_FIXES[t]
#     return t


# def is_col_header(tok: str) -> bool:
#     t = fix_ocr_token(tok)
#     if t in GROUP_HEADERS or t in ROW_LABELS:
#         return False
#     if t in KNOWN_COLS:
#         return True
#     # Accept any 2–8 uppercase letter sequence not obviously noise
#     return bool(re.fullmatch(r"[A-Z+]{2,8}", t))


# # ─────────────────────────────────────────────────────────────────────────────
# # Layout detection
# # ─────────────────────────────────────────────────────────────────────────────

# def find_dark_separator(arr: "np.ndarray", start_frac: float = 0.35) -> int | None:
#     """
#     Find the single thick black horizontal line that separates the bar chart
#     from the data table.  Returns the y-coordinate or None.
#     """
#     import numpy as np
#     h, w = arr.shape[:2]
#     for y in range(int(h * start_frac), h):
#         dark = int(np.sum((arr[y, :, 0] < 80) & (arr[y, :, 1] < 80) & (arr[y, :, 2] < 80)))
#         if dark > w * 0.75:
#             return y
#     return None


# def find_score_total_y(img, sep_y: int, w: int):
#     """
#     Search for the "Score" label in the left margin to pinpoint exact row y.
#     Returns (score_y, total_y) as floats.
#     """
#     pytesseract = setup_tesseract()
#     search_top  = sep_y + 40
#     search_bot  = sep_y + 200
#     # Only scan the left ~200px where the label lives
#     region = img.crop((0, search_top, min(200, w), search_bot))
#     scale  = 4
#     region_big = region.resize((region.size[0] * scale, region.size[1] * scale), _lanczos())
#     data = pytesseract.image_to_data(
#         region_big.convert("L"),
#         config="--psm 11",
#         output_type=pytesseract.Output.DICT,
#     )
#     score_y = total_y = None
#     for i, text in enumerate(data["text"]):
#         t = text.strip().lower()
#         if "score" in t and score_y is None:
#             score_y = (data["top"][i] + data["height"][i] / 2) / scale + search_top
#         if "total" in t and total_y is None:
#             total_y = (data["top"][i] + data["height"][i] / 2) / scale + search_top

#     if score_y is None:
#         score_y = sep_y + 80
#     if total_y is None:
#         total_y = score_y + 32

#     return score_y, total_y


# def _lanczos():
#     from PIL import Image
#     return Image.LANCZOS


# # ─────────────────────────────────────────────────────────────────────────────
# # Column header OCR — three-pass
# # ─────────────────────────────────────────────────────────────────────────────

# def ocr_full_header_band(pytesseract, img, sep_y: int, w: int, scale: int = 4):
#     """Pass A: full-band PSM 11 on the column header strip."""
#     from PIL import Image
#     crop = img.crop((0, sep_y - 38, w, sep_y))
#     big  = crop.resize((w * scale, crop.size[1] * scale), Image.LANCZOS)
#     gray = big.convert("L")
#     data = pytesseract.image_to_data(
#         gray,
#         config="--psm 11 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ+",
#         output_type=pytesseract.Output.DICT,
#     )
#     items = []
#     for i, text in enumerate(data["text"]):
#         text = text.strip()
#         if not text:
#             continue
#         conf  = data["conf"][i]
#         fixed = fix_ocr_token(text)
#         # Tesseract gives conf=0 for multi-word tokens like "OE+MIR" in some modes.
#         # Accept any non-negative confidence if the token resolves to a known column;
#         # otherwise keep the conf>5 gate to filter random noise.
#         if conf < 0:
#             continue
#         if conf < 5 and fixed not in KNOWN_COLS:
#             continue
#         xc = (data["left"][i] + data["width"][i] / 2) / scale
#         items.append((xc, text))
#     items.sort(key=lambda t: t[0])
#     return items


# def ocr_per_column_header(pytesseract, img, xc: float, sep_y: int, w: int, scale: int = 6):
#     """Pass B: narrow crop centred on one column, high upscale."""
#     from PIL import Image
#     margin = 48
#     x1 = max(0, int(xc - margin))
#     x2 = min(w, int(xc + margin))
#     crop = img.crop((x1, sep_y - 38, x2, sep_y))
#     big  = crop.resize((crop.size[0] * scale, crop.size[1] * scale), Image.LANCZOS)
#     gray = big.convert("L")
#     text = pytesseract.image_to_string(
#         gray,
#         config="--psm 6 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ+",
#     ).strip().replace("\n", " ").strip()
#     # Take the longest token
#     tokens = [t for t in text.split() if t]
#     return max(tokens, key=len) if tokens else ""


# def build_header_map(pytesseract, img, sep_y: int, w: int, score_xs: list[float]) -> dict[float, str]:
#     """
#     Return {score_x: col_name} using three-pass OCR + fuzzy correction.
#     score_xs are the x-centres of score values (reliable ground truth for column positions).
#     """

#     # ── Pass A ───────────────────────────────────────────────────────────────
#     raw_items = ocr_full_header_band(pytesseract, img, sep_y, w)
#     # Filter to plausible column headers
#     pass_a = [(x, fix_ocr_token(t)) for x, t in raw_items if is_col_header(t)]

#     # ── Map Pass A results to score_xs by nearest neighbour ──────────────────
#     assigned: dict[int, str] = {}  # score_xs index → col name
#     used_score = set()
#     for hx, hname in pass_a:
#         best_dist, best_idx = float("inf"), None
#         for i, sx in enumerate(score_xs):
#             if i in used_score:
#                 continue
#             d = abs(hx - sx)
#             if d < best_dist:
#                 best_dist, best_idx = d, i
#         if best_idx is not None and best_dist < 120:
#             assigned[best_idx] = hname
#             used_score.add(best_idx)

#     # ── Pass B: fill gaps ─────────────────────────────────────────────────────
#     missing = [i for i in range(len(score_xs)) if i not in assigned]
#     for i in missing:
#         xc = score_xs[i]
#         raw = ocr_per_column_header(pytesseract, img, xc, sep_y, w)
#         if raw:
#             fixed = fix_ocr_token(raw)
#             # Try prefix completion if still not in KNOWN_COLS
#             if fixed not in KNOWN_COLS and fixed in PREFIX_COMPLETIONS:
#                 fixed = PREFIX_COMPLETIONS[fixed]
#             if is_col_header(fixed):
#                 assigned[i] = fixed

#     # ── Pass C: prefix-complete any truncated tokens ──────────────────────────
#     for i, name in list(assigned.items()):
#         if name not in KNOWN_COLS and name in PREFIX_COMPLETIONS:
#             assigned[i] = PREFIX_COMPLETIONS[name]

#     # Build final dict keyed by score_x
#     return {score_xs[i]: assigned.get(i, f"COL{i+1}") for i in range(len(score_xs))}


# # ─────────────────────────────────────────────────────────────────────────────
# # Value extraction
# # ─────────────────────────────────────────────────────────────────────────────

# def extract_values_at_y(pytesseract, img, y_centre: float, w: int, scale: int = 4):
#     """
#     Extract (x_centre, value_string) pairs from a horizontal data row.
#     Returns floats only.
#     """
#     import numpy as np
#     from PIL import Image

#     half = 14
#     y1, y2 = max(0, int(y_centre - half)), int(y_centre + half)
#     crop = img.crop((80, y1, w, y2))    # skip left label area
#     big  = crop.resize((crop.size[0] * scale, crop.size[1] * scale), Image.LANCZOS)

#     arr  = np.array(big)
#     # Max-channel binarize to strip bar-chart bleed
#     max_ch = np.maximum(np.maximum(arr[:,:,0].astype(int), arr[:,:,1].astype(int)), arr[:,:,2].astype(int))
#     clean  = np.where(max_ch < 160, 0, 255).astype("uint8")
#     from PIL import Image as PILImage
#     pil_clean = PILImage.fromarray(clean)

#     data = pytesseract.image_to_data(
#         pil_clean,
#         config="--psm 11 -c tessedit_char_whitelist=.0123456789",
#         output_type=pytesseract.Output.DICT,
#     )
#     items = []
#     for i, text in enumerate(data["text"]):
#         text = text.strip()
#         if not text or data["conf"][i] < 5:
#             continue
#         xc = (data["left"][i] + data["width"][i] / 2) / scale + 80
#         items.append((xc, text))
#     items.sort(key=lambda t: t[0])
#     return items


# def smart_fix_decimal(s: str) -> str:
#     """
#     Recover a missing decimal point: '1391' → '13.91', '2765' → '27.65'.
#     NIRF values always have exactly 2 decimal places.
#     """
#     s = s.strip()
#     if not s:
#         return s
#     if re.fullmatch(r"\d+\.\d+", s):
#         return s   # already correct
#     if s.isdigit() and len(s) >= 3:
#         return s[:-2] + "." + s[-2:]
#     return s


# def is_float(s: str) -> bool:
#     try:
#         float(s)
#         return True
#     except ValueError:
#         return False


# def match_values_to_headers(
#     header_map: dict[float, str],
#     value_items: list[tuple[float, str]],
# ) -> dict[str, str]:
#     """Nearest-neighbour match between header x-positions and value x-positions."""
#     result:  dict[str, str] = {}
#     used = set()
#     for hx, col_name in sorted(header_map.items()):
#         best_dist, best_val = float("inf"), ""
#         for i, (vx, vtext) in enumerate(value_items):
#             if i in used:
#                 continue
#             d = abs(hx - vx)
#             if d < best_dist:
#                 best_dist, best_val = d, vtext
#                 best_idx = i
#         if best_dist < 100:
#             used.add(best_idx)
#             result[col_name] = smart_fix_decimal(best_val)
#     return result


# # ─────────────────────────────────────────────────────────────────────────────
# # Institute name / code
# # ─────────────────────────────────────────────────────────────────────────────

# def extract_institute_meta(pytesseract, img, arr: "np.ndarray"):
#     import numpy as np
#     from PIL import Image
#     h, w = arr.shape[:2]

#     # Code from filename (always reliable)
#     fname = Path(getattr(img, "filename", "")).stem
#     institute_code = fname if re.match(r"IR-[A-Z]-[A-Z]-\d+", fname) else ""

#     # Find title rows: scan top 90 px for non-white content
#     t_start = t_end = None
#     for y in range(min(90, h)):
#         nw = int(np.sum(~((arr[y,:,0] > 210) & (arr[y,:,1] > 210) & (arr[y,:,2] > 210))))
#         if nw > 30:
#             if t_start is None:
#                 t_start = y
#             t_end = y

#     if t_start is None:
#         t_start, t_end = 5, 30

#     title_crop = img.crop((0, max(0, t_start - 4), w, min(h, t_end + 6)))
#     tw, th = title_crop.size
#     title_big = title_crop.resize((tw * 5, th * 5), Image.LANCZOS)
#     title_text = pytesseract.image_to_string(title_big.convert("L"), config="--psm 7").strip()

#     m = re.search(r"(.+?)\s*\(IR-[A-Z]-[A-Z]-\d+\)", title_text)
#     if m:
#         institute_name = m.group(1).strip()
#     else:
#         lines = [l.strip() for l in title_text.splitlines() if len(l.strip()) > 4]
#         institute_name = max(lines, key=len) if lines else ""
#     institute_name = re.sub(r"^[^A-Za-z;]+", "", institute_name).strip().lstrip(";").strip()

#     return institute_name, institute_code


# # ─────────────────────────────────────────────────────────────────────────────
# # Per-image extraction entry point
# # ─────────────────────────────────────────────────────────────────────────────

# def extract_one(image_path: str) -> dict:
#     from PIL import Image
#     import numpy as np

#     pytesseract = setup_tesseract()

#     img = Image.open(image_path).convert("RGB")
#     img.filename = image_path   # keep for meta extraction
#     arr = np.array(img)
#     h, w = arr.shape[:2]

#     # ── Institute meta ────────────────────────────────────────────────────────
#     institute_name, institute_code = extract_institute_meta(pytesseract, img, arr)

#     # ── Find layout anchor ────────────────────────────────────────────────────
#     sep_y = find_dark_separator(arr)
#     if sep_y is None:
#         # Graceful degradation: no separator found
#         return {
#             "image_path": image_path, "year": "", "category": "",
#             "institute_name": institute_name, "institute_code": institute_code,
#             "table": {},
#             "_msg": f"  NO_SEP {institute_code} | {institute_name[:40]}",
#         }

#     # ── Find Score / Total row y-positions ───────────────────────────────────
#     score_y, total_y = find_score_total_y(img, sep_y, w)

#     # ── Extract score values (ground-truth x-positions) ───────────────────────
#     score_items = extract_values_at_y(pytesseract, img, score_y, w)
#     score_float = [(x, smart_fix_decimal(v)) for x, v in score_items if is_float(smart_fix_decimal(v))]
#     score_xs = [x for x, _ in score_float]

#     if not score_xs:
#         return {
#             "image_path": image_path, "year": "", "category": "",
#             "institute_name": institute_name, "institute_code": institute_code,
#             "table": {},
#             "_msg": f"  NO_SCORES {institute_code} | {institute_name[:40]}",
#         }

#     # ── Build column header map ───────────────────────────────────────────────
#     header_map = build_header_map(pytesseract, img, sep_y, w, score_xs)

#     # ── Extract total values ───────────────────────────────────────────────────
#     total_items = extract_values_at_y(pytesseract, img, total_y, w)
#     total_float = [(x, smart_fix_decimal(v)) for x, v in total_items if is_float(smart_fix_decimal(v))]

#     # ── Match values to headers ───────────────────────────────────────────────
#     score_map = match_values_to_headers(header_map, score_float)
#     total_map = match_values_to_headers(header_map, total_float)

#     table = {
#         col_name: {
#             "score": score_map.get(col_name, ""),
#             "total": total_map.get(col_name, ""),
#         }
#         for col_name in header_map.values()
#     }

#     # ── Year / category from folder structure ─────────────────────────────────
#     parts = Path(image_path).resolve().parts
#     year = category = ""
#     for i, p in enumerate(parts):
#         if p == "image" and i >= 2:
#             year, category = parts[i - 2], parts[i - 1]
#             break

#     n_cols   = len(table)
#     n_scores = sum(1 for v in table.values() if v["score"])
#     flag     = " MISMATCH" if n_cols > 0 and n_cols != n_scores else " OK"

#     return {
#         "image_path":     image_path,
#         "year":           year,
#         "category":       category,
#         "institute_name": institute_name,
#         "institute_code": institute_code,
#         "table":          table,
#         "_msg":           (
#             f"  {institute_code} | {(institute_name or '?')[:40]} "
#             f"| cols={n_cols} scores={n_scores}{flag}"
#         ),
#     }


# # ─────────────────────────────────────────────────────────────────────────────
# # Image collection
# # ─────────────────────────────────────────────────────────────────────────────

# IMAGE_EXTS = {".jpg", ".jpeg", ".png"}


# def collect_images(path: str) -> list[str]:
#     p = Path(os.path.normpath(path))
#     if p.is_file() and p.suffix.lower() in IMAGE_EXTS:
#         return [str(p)]
#     if p.is_dir():
#         imgs = []
#         for ext in IMAGE_EXTS:
#             imgs.extend(str(f) for f in p.rglob(f"*{ext}"))
#         imgs.sort()
#         return imgs
#     print(f"Path not found: {p}", file=sys.stderr)
#     return []


# # ─────────────────────────────────────────────────────────────────────────────
# # Excel output
# # ─────────────────────────────────────────────────────────────────────────────

# def build_excel(rows: list[dict], output_path: str) -> None:
#     from openpyxl import Workbook
#     from openpyxl.styles import Font, PatternFill, Alignment

#     header_fill  = PatternFill("solid", fgColor="1F4E79")
#     header_font  = Font(color="FFFFFF", bold=True)
#     alt_fill     = PatternFill("solid", fgColor="D9E1F2")
#     center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

#     all_headers: list[str] = []
#     seen_h: set[str] = set()
#     for row in rows:
#         for hk in row.get("table", {}):
#             if hk not in seen_h:
#                 seen_h.add(hk)
#                 all_headers.append(hk)

#     fixed_cols = ["Year", "Category", "Institute Name", "Institute Code"]
#     all_cols   = (
#         fixed_cols
#         + [f"{h} Score" for h in all_headers]
#         + [f"{h} Total" for h in all_headers]
#     )

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "NIRF Scores"

#     for ci, col in enumerate(all_cols, 1):
#         cell = ws.cell(row=1, column=ci, value=col)
#         cell.fill  = header_fill
#         cell.font  = header_font
#         cell.alignment = center_align
#     ws.row_dimensions[1].height = 30

#     for ri, row in enumerate(rows, 2):
#         tbl  = row.get("table", {})
#         fill = alt_fill if ri % 2 == 0 else PatternFill()
#         vals = (
#             [row.get("year",""), row.get("category",""),
#              row.get("institute_name",""), row.get("institute_code","")]
#             + [tbl.get(h, {}).get("score", "") for h in all_headers]
#             + [tbl.get(h, {}).get("total", "") for h in all_headers]
#         )
#         for ci, val in enumerate(vals, 1):
#             cell = ws.cell(row=ri, column=ci, value=val)
#             cell.fill      = fill
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#     for ci in range(1, len(all_cols) + 1):
#         col_letter = ws.cell(row=1, column=ci).column_letter
#         max_len = max(
#             len(str(ws.cell(row=1, column=ci).value or "")),
#             *(len(str(ws.cell(row=ri, column=ci).value or "")) for ri in range(2, len(rows) + 2)),
#         )
#         ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

#     ws.freeze_panes = "A2"
#     wb.save(output_path)
#     print(f"\nDone. Saved {len(rows)} rows → {output_path}", flush=True)


# # ─────────────────────────────────────────────────────────────────────────────
# # Main
# # ─────────────────────────────────────────────────────────────────────────────

# def main() -> None:
#     parser = argparse.ArgumentParser(description="Extract NIRF score-card table data (v4)")
#     parser.add_argument("input",     help="Image file or folder (searched recursively)")
#     parser.add_argument("--output",  "-o", default="nirf_scores.xlsx", help="Output Excel file")
#     parser.add_argument("--json",    "-j", default=None,               help="Also write JSON output")
#     parser.add_argument(
#         "--workers", "-w",
#         type=int,
#         default=min(8, multiprocessing.cpu_count()),
#         help="Parallel threads (default: min(8, cpu_count))",
#     )
#     args = parser.parse_args()

#     images = collect_images(args.input)
#     if not images:
#         print(f"No images found at: {args.input}", file=sys.stderr)
#         sys.exit(1)

#     n = len(images)
#     w = min(args.workers, n)
#     print(f"Found {n} image(s).  Threads: {w}\n", flush=True)

#     # Quick sanity-check for Tesseract
#     try:
#         pyt = setup_tesseract()
#         from PIL import Image as PILImage
#         import numpy as np
#         pyt.image_to_string(PILImage.fromarray(np.ones((10, 10, 3), dtype=np.uint8) * 255))
#         print("Tesseract OK.\n", flush=True)
#     except Exception as e:
#         print(f"\nERROR: Tesseract not available: {e}", file=sys.stderr)
#         print("Install from: https://github.com/UB-Mannheim/tesseract/wiki", file=sys.stderr)
#         sys.exit(1)

#     rows_by_path: dict[str, dict] = {}
#     done = 0

#     with ThreadPoolExecutor(max_workers=w) as exe:
#         futures = {exe.submit(extract_one, p): p for p in images}
#         for fut in as_completed(futures):
#             done += 1
#             img_path = futures[fut]
#             try:
#                 row = fut.result()
#                 msg = row.pop("_msg", "")
#                 rows_by_path[img_path] = row
#                 print(f"[{done}/{n}] {Path(img_path).name}{msg}", flush=True)
#             except Exception as e:
#                 print(f"[{done}/{n}] ERROR {Path(img_path).name}: {e}", flush=True)

#     rows = [rows_by_path[p] for p in images if p in rows_by_path]

#     if args.json:
#         with open(args.json, "w", encoding="utf-8") as f:
#             json.dump(rows, f, indent=2, ensure_ascii=False)
#         print(f"JSON → {args.json}", flush=True)

#     if not rows:
#         print("No rows extracted.", flush=True)
#         sys.exit(0)

#     build_excel(rows, args.output)


# if __name__ == "__main__":
#     main()














































#WORKING FOR 2019 ALSO








# # -*- coding: utf-8 -*-
# """
# paddle_extract.py  —  NIRF score-card image → Excel extractor (v4)
# """

# import io, os, sys, re, json, argparse
# from pathlib import Path
# from concurrent.futures import ThreadPoolExecutor, as_completed
# import multiprocessing

# if hasattr(sys.stdout, "reconfigure"):
#     sys.stdout.reconfigure(encoding="utf-8", errors="replace")
# else:
#     sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# TESSERACT_CMD = None   # set to r"C:\Program Files\Tesseract-OCR\tesseract.exe" if needed

# # ─────────────────────────────────────────────────────────────────────────────
# # Tesseract setup
# # ─────────────────────────────────────────────────────────────────────────────

# def setup_tesseract():
#     import pytesseract
#     if TESSERACT_CMD:
#         pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
#     else:
#         default = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
#         if os.path.exists(default):
#             pytesseract.pytesseract.tesseract_cmd = default
#     return pytesseract


# # ─────────────────────────────────────────────────────────────────────────────
# # Known NIRF column metadata
# # ─────────────────────────────────────────────────────────────────────────────

# KNOWN_COLS = {
#     "SS", "FSR", "FQE", "FRU",
#     "OE+MIR", "OE", "MIR",
#     "PU", "QP", "IPR", "FPPP", "SDG",
#     "GPH", "GUE", "MS", "GPHD",
#     "RD", "WD", "ESCS", "PCS",
#     "PR", "PERCEPTION",
# }

# GROUP_HEADERS = {"TLR", "RP", "GO", "OI"}
# ROW_LABELS    = {"SCORE", "TOTAL", "RANK"}

# OCR_FIXES = {
#     "S5": "SS", "5S": "SS", "S$": "SS", "55": "SS",
#     "W0": "WD", "VVD": "WD", "WO": "WD",
#     "FQF": "FQE", "FOE": "FQE", "FDE": "FQE",
#     "FRO": "FRU", "FR0": "FRU",
#     "1PR": "IPR", "lPR": "IPR",
#     "GU3": "GUE",
#     "P5": "PCS", "PC5": "PCS",
#     "E5CS": "ESCS", "ESC5": "ESCS", "ECSS": "ESCS",
#     "0E+MIR": "OE+MIR", "OE-MIR": "OE+MIR", "OE+MR": "OE+MIR",
#     "5DG": "SDG", "SDC": "SDG",
#     "M5": "MS", "GP H": "GPH",
#     "OP": "QP", "OF": "OE",
#     "GPHD": "GPHD", "GPHO": "GPHD",
#     "SR": "FSR", "FSP": "FSR",
#     "OE": "OE", "MlR": "MIR",
#     "FRl": "FRU", "FRI": "FRU",
# }

# PREFIX_COMPLETIONS = {
#     "SR": "FSR", "QE": "FQE", "RU": "FRU",
#     "S": "SS", "PH": "GPH", "UE": "GUE",
#     "HD": "GPHD", "CS": "PCS", "SCS": "ESCS",
# }


# def fix_ocr_token(tok: str) -> str:
#     t = tok.upper().replace(" ", "")
#     if t in OCR_FIXES:
#         return OCR_FIXES[t]
#     return t


# def is_col_header(tok: str) -> bool:
#     t = fix_ocr_token(tok)
#     if t in GROUP_HEADERS or t in ROW_LABELS:
#         return False
#     if t in KNOWN_COLS:
#         return True
#     return bool(re.fullmatch(r"[A-Z+]{2,8}", t))


# # ─────────────────────────────────────────────────────────────────────────────
# # Layout detection
# # ─────────────────────────────────────────────────────────────────────────────

# def find_dark_separator(arr, start_frac=0.35):
#     import numpy as np
#     h, w = arr.shape[:2]
#     for y in range(int(h * start_frac), h):
#         dark = int(np.sum((arr[y, :, 0] < 80) & (arr[y, :, 1] < 80) & (arr[y, :, 2] < 80)))
#         if dark > w * 0.75:
#             return y
#     return None


# def find_score_total_y(img, sep_y, w):
#     pytesseract = setup_tesseract()
#     search_top  = sep_y + 40
#     search_bot  = sep_y + 200
#     region = img.crop((0, search_top, min(200, w), search_bot))
#     scale  = 4
#     region_big = region.resize((region.size[0] * scale, region.size[1] * scale), _lanczos())
#     data = pytesseract.image_to_data(
#         region_big.convert("L"),
#         config="--psm 11",
#         output_type=pytesseract.Output.DICT,
#     )
#     score_y = total_y = None
#     for i, text in enumerate(data["text"]):
#         t = text.strip().lower()
#         if "score" in t and score_y is None:
#             score_y = (data["top"][i] + data["height"][i] / 2) / scale + search_top
#         if "total" in t and total_y is None:
#             total_y = (data["top"][i] + data["height"][i] / 2) / scale + search_top

#     if score_y is None:
#         score_y = sep_y + 80
#     if total_y is None:
#         total_y = score_y + 32

#     return score_y, total_y


# def _lanczos():
#     from PIL import Image
#     return Image.LANCZOS


# # ─────────────────────────────────────────────────────────────────────────────
# # Column header OCR — three-pass
# # ─────────────────────────────────────────────────────────────────────────────

# def ocr_full_header_band(pytesseract, img, sep_y, w, scale=4):
#     from PIL import Image
#     crop = img.crop((0, sep_y - 38, w, sep_y))
#     big  = crop.resize((w * scale, crop.size[1] * scale), Image.LANCZOS)
#     gray = big.convert("L")
#     data = pytesseract.image_to_data(
#         gray,
#         config="--psm 11 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ+",
#         output_type=pytesseract.Output.DICT,
#     )
#     items = []
#     for i, text in enumerate(data["text"]):
#         text = text.strip()
#         if not text:
#             continue
#         conf  = data["conf"][i]
#         fixed = fix_ocr_token(text)
#         if conf < 0:
#             continue
#         if conf < 5 and fixed not in KNOWN_COLS:
#             continue
#         xc = (data["left"][i] + data["width"][i] / 2) / scale
#         items.append((xc, text))
#     items.sort(key=lambda t: t[0])
#     return items


# def ocr_per_column_header(pytesseract, img, xc, sep_y, w, scale=6):
#     from PIL import Image
#     margin = 48
#     x1 = max(0, int(xc - margin))
#     x2 = min(w, int(xc + margin))
#     crop = img.crop((x1, sep_y - 38, x2, sep_y))
#     big  = crop.resize((crop.size[0] * scale, crop.size[1] * scale), Image.LANCZOS)
#     gray = big.convert("L")
#     text = pytesseract.image_to_string(
#         gray,
#         config="--psm 6 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ+",
#     ).strip().replace("\n", " ").strip()
#     tokens = [t for t in text.split() if t]
#     return max(tokens, key=len) if tokens else ""


# def build_header_map(pytesseract, img, sep_y, w, score_xs):
#     raw_items = ocr_full_header_band(pytesseract, img, sep_y, w)
#     pass_a = [(x, fix_ocr_token(t)) for x, t in raw_items if is_col_header(t)]

#     assigned = {}
#     used_score = set()
#     for hx, hname in pass_a:
#         best_dist, best_idx = float("inf"), None
#         for i, sx in enumerate(score_xs):
#             if i in used_score:
#                 continue
#             d = abs(hx - sx)
#             if d < best_dist:
#                 best_dist, best_idx = d, i
#         if best_idx is not None and best_dist < 120:
#             assigned[best_idx] = hname
#             used_score.add(best_idx)

#     missing = [i for i in range(len(score_xs)) if i not in assigned]
#     for i in missing:
#         xc = score_xs[i]
#         raw = ocr_per_column_header(pytesseract, img, xc, sep_y, w)
#         if raw:
#             fixed = fix_ocr_token(raw)
#             if fixed not in KNOWN_COLS and fixed in PREFIX_COMPLETIONS:
#                 fixed = PREFIX_COMPLETIONS[fixed]
#             if is_col_header(fixed):
#                 assigned[i] = fixed

#     for i, name in list(assigned.items()):
#         if name not in KNOWN_COLS and name in PREFIX_COMPLETIONS:
#             assigned[i] = PREFIX_COMPLETIONS[name]

#     return {score_xs[i]: assigned.get(i, f"COL{i+1}") for i in range(len(score_xs))}


# # ─────────────────────────────────────────────────────────────────────────────
# # Value extraction
# # ─────────────────────────────────────────────────────────────────────────────

# def extract_values_at_y(pytesseract, img, y_centre, w, scale=4):
#     import numpy as np
#     from PIL import Image

#     half = 14
#     y1, y2 = max(0, int(y_centre - half)), int(y_centre + half)
#     crop = img.crop((80, y1, w, y2))
#     big  = crop.resize((crop.size[0] * scale, crop.size[1] * scale), Image.LANCZOS)

#     arr    = np.array(big)
#     max_ch = np.maximum(np.maximum(arr[:,:,0].astype(int), arr[:,:,1].astype(int)), arr[:,:,2].astype(int))
#     clean  = np.where(max_ch < 160, 0, 255).astype("uint8")
#     from PIL import Image as PILImage
#     pil_clean = PILImage.fromarray(clean)

#     data = pytesseract.image_to_data(
#         pil_clean,
#         config="--psm 11 -c tessedit_char_whitelist=.0123456789",
#         output_type=pytesseract.Output.DICT,
#     )
#     items = []
#     for i, text in enumerate(data["text"]):
#         text = text.strip()
#         if not text or data["conf"][i] < 5:
#             continue
#         xc = (data["left"][i] + data["width"][i] / 2) / scale + 80
#         items.append((xc, text))
#     items.sort(key=lambda t: t[0])
#     return items


# def smart_fix_decimal(s):
#     s = s.strip()
#     if not s:
#         return s
#     if re.fullmatch(r"\d+\.\d+", s):
#         return s
#     if s.isdigit() and len(s) >= 3:
#         return s[:-2] + "." + s[-2:]
#     return s


# def is_float(s):
#     try:
#         float(s)
#         return True
#     except ValueError:
#         return False


# def match_values_to_headers(header_map, value_items):
#     result = {}
#     used = set()
#     for hx, col_name in sorted(header_map.items()):
#         best_dist, best_val = float("inf"), ""
#         best_idx = None
#         for i, (vx, vtext) in enumerate(value_items):
#             if i in used:
#                 continue
#             d = abs(hx - vx)
#             if d < best_dist:
#                 best_dist, best_val = d, vtext
#                 best_idx = i
#         if best_dist < 100 and best_idx is not None:
#             used.add(best_idx)
#             result[col_name] = smart_fix_decimal(best_val)
#     return result


# # ─────────────────────────────────────────────────────────────────────────────
# # Institute name / code  — THREE SOURCES (fixes missing code for 2019 images)
# # ─────────────────────────────────────────────────────────────────────────────

# def extract_institute_meta(pytesseract, img, arr):
#     import numpy as np
#     from PIL import Image
#     h, w = arr.shape[:2]

#     # ── Source 1: filename (works for 2020+ images named IR-X-X-XXXX) ────────
#     fname = Path(getattr(img, "filename", "")).stem
#     institute_code = fname if re.match(r"IR-[A-Z]-[A-Z]-\d+", fname) else ""

#     # ── Find title rows: scan top 90px for non-white content ─────────────────
#     t_start = t_end = None
#     for y in range(min(90, h)):
#         nw = int(np.sum(~((arr[y,:,0] > 210) & (arr[y,:,1] > 210) & (arr[y,:,2] > 210))))
#         if nw > 30:
#             if t_start is None:
#                 t_start = y
#             t_end = y

#     if t_start is None:
#         t_start, t_end = 5, 30

#     title_crop = img.crop((0, max(0, t_start - 4), w, min(h, t_end + 6)))
#     tw, th = title_crop.size
#     title_big = title_crop.resize((tw * 5, th * 5), Image.LANCZOS)
#     title_text = pytesseract.image_to_string(title_big.convert("L"), config="--psm 7").strip()

#     # ── Source 2: parse code from OCR title text e.g. "(IR-O-U-0391)" ────────
#     # FIX: 2019 image filenames don't follow IR-X-X-XXXX format so Source 1
#     # returns "". Fall back to reading the code from the image title text itself.
#     if not institute_code:
#         m_code = re.search(r"\(?(IR-[A-Z]-[A-Z]-\d+)\)?", title_text)
#         if m_code:
#             institute_code = m_code.group(1).strip()

#     # ── Source 3: wider PSM-11 scan of top 20% of image ──────────────────────
#     # Last resort — in case code appears on a different line in the header area
#     if not institute_code:
#         header_crop = img.crop((0, 0, w, int(h * 0.20)))
#         header_big  = header_crop.resize((w * 3, header_crop.size[1] * 3), Image.LANCZOS)
#         header_text = pytesseract.image_to_string(
#             header_big.convert("L"), config="--psm 11"
#         ).strip()
#         m_code2 = re.search(r"IR-[A-Z]-[A-Z]-\d+", header_text)
#         if m_code2:
#             institute_code = m_code2.group(0).strip()

#     # ── Extract institute name ────────────────────────────────────────────────
#     m = re.search(r"(.+?)\s*\(IR-[A-Z]-[A-Z]-\d+\)", title_text)
#     if m:
#         institute_name = m.group(1).strip()
#     else:
#         lines = [l.strip() for l in title_text.splitlines() if len(l.strip()) > 4]
#         institute_name = max(lines, key=len) if lines else ""
#     institute_name = re.sub(r"^[^A-Za-z;]+", "", institute_name).strip().lstrip(";").strip()

#     return institute_name, institute_code


# # ─────────────────────────────────────────────────────────────────────────────
# # Per-image extraction entry point
# # ─────────────────────────────────────────────────────────────────────────────

# def extract_one(image_path):
#     from PIL import Image
#     import numpy as np

#     pytesseract = setup_tesseract()

#     img = Image.open(image_path).convert("RGB")
#     img.filename = image_path
#     arr = np.array(img)
#     h, w = arr.shape[:2]

#     institute_name, institute_code = extract_institute_meta(pytesseract, img, arr)

#     sep_y = find_dark_separator(arr)
#     if sep_y is None:
#         return {
#             "image_path": image_path, "year": "", "category": "",
#             "institute_name": institute_name, "institute_code": institute_code,
#             "table": {},
#             "_msg": f"  NO_SEP {institute_code} | {institute_name[:40]}",
#         }

#     score_y, total_y = find_score_total_y(img, sep_y, w)

#     score_items = extract_values_at_y(pytesseract, img, score_y, w)
#     score_float = [(x, smart_fix_decimal(v)) for x, v in score_items if is_float(smart_fix_decimal(v))]
#     score_xs = [x for x, _ in score_float]

#     if not score_xs:
#         return {
#             "image_path": image_path, "year": "", "category": "",
#             "institute_name": institute_name, "institute_code": institute_code,
#             "table": {},
#             "_msg": f"  NO_SCORES {institute_code} | {institute_name[:40]}",
#         }

#     header_map = build_header_map(pytesseract, img, sep_y, w, score_xs)

#     total_items = extract_values_at_y(pytesseract, img, total_y, w)
#     total_float = [(x, smart_fix_decimal(v)) for x, v in total_items if is_float(smart_fix_decimal(v))]

#     score_map = match_values_to_headers(header_map, score_float)
#     total_map = match_values_to_headers(header_map, total_float)

#     table = {
#         col_name: {
#             "score": score_map.get(col_name, ""),
#             "total": total_map.get(col_name, ""),
#         }
#         for col_name in header_map.values()
#     }

#     parts = Path(image_path).resolve().parts
#     year = category = ""
#     for i, p in enumerate(parts):
#         if p == "image" and i >= 2:
#             year, category = parts[i - 2], parts[i - 1]
#             break

#     n_cols   = len(table)
#     n_scores = sum(1 for v in table.values() if v["score"])
#     flag     = " MISMATCH" if n_cols > 0 and n_cols != n_scores else " OK"

#     return {
#         "image_path":     image_path,
#         "year":           year,
#         "category":       category,
#         "institute_name": institute_name,
#         "institute_code": institute_code,
#         "table":          table,
#         "_msg": (
#             f"  {institute_code} | {(institute_name or '?')[:40]} "
#             f"| cols={n_cols} scores={n_scores}{flag}"
#         ),
#     }


# # ─────────────────────────────────────────────────────────────────────────────
# # Image collection
# # ─────────────────────────────────────────────────────────────────────────────

# IMAGE_EXTS = {".jpg", ".jpeg", ".png"}


# def collect_images(path):
#     p = Path(os.path.normpath(path))
#     if p.is_file() and p.suffix.lower() in IMAGE_EXTS:
#         return [str(p)]
#     if p.is_dir():
#         imgs = []
#         for ext in IMAGE_EXTS:
#             imgs.extend(str(f) for f in p.rglob(f"*{ext}"))
#         imgs.sort()
#         return imgs
#     print(f"Path not found: {p}", file=sys.stderr)
#     return []


# # ─────────────────────────────────────────────────────────────────────────────
# # Excel output
# # ─────────────────────────────────────────────────────────────────────────────

# def build_excel(rows, output_path):
#     from openpyxl import Workbook
#     from openpyxl.styles import Font, PatternFill, Alignment

#     header_fill  = PatternFill("solid", fgColor="1F4E79")
#     header_font  = Font(color="FFFFFF", bold=True)
#     alt_fill     = PatternFill("solid", fgColor="D9E1F2")
#     center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

#     all_headers = []
#     seen_h = set()
#     for row in rows:
#         for hk in row.get("table", {}):
#             if hk not in seen_h:
#                 seen_h.add(hk)
#                 all_headers.append(hk)

#     fixed_cols = ["Year", "Category", "Institute Name", "Institute Code"]
#     all_cols   = (
#         fixed_cols
#         + [f"{h} Score" for h in all_headers]
#         + [f"{h} Total" for h in all_headers]
#     )

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "NIRF Scores"

#     for ci, col in enumerate(all_cols, 1):
#         cell = ws.cell(row=1, column=ci, value=col)
#         cell.fill      = header_fill
#         cell.font      = header_font
#         cell.alignment = center_align
#     ws.row_dimensions[1].height = 30

#     for ri, row in enumerate(rows, 2):
#         tbl  = row.get("table", {})
#         fill = alt_fill if ri % 2 == 0 else PatternFill()
#         vals = (
#             [row.get("year",""), row.get("category",""),
#              row.get("institute_name",""), row.get("institute_code","")]
#             + [tbl.get(h, {}).get("score", "") for h in all_headers]
#             + [tbl.get(h, {}).get("total", "") for h in all_headers]
#         )
#         for ci, val in enumerate(vals, 1):
#             cell = ws.cell(row=ri, column=ci, value=val)
#             cell.fill      = fill
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#     for ci in range(1, len(all_cols) + 1):
#         col_letter = ws.cell(row=1, column=ci).column_letter
#         max_len = max(
#             len(str(ws.cell(row=1, column=ci).value or "")),
#             *(len(str(ws.cell(row=ri, column=ci).value or "")) for ri in range(2, len(rows) + 2)),
#         )
#         ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

#     ws.freeze_panes = "A2"
#     wb.save(output_path)
#     print(f"\nDone. Saved {len(rows)} rows → {output_path}", flush=True)


# # ─────────────────────────────────────────────────────────────────────────────
# # Main
# # ─────────────────────────────────────────────────────────────────────────────

# def main():
#     parser = argparse.ArgumentParser(description="Extract NIRF score-card table data (v4)")
#     parser.add_argument("input",    help="Image file or folder (searched recursively)")
#     parser.add_argument("--output", "-o", default="nirf_scores.xlsx", help="Output Excel file")
#     parser.add_argument("--json",   "-j", default=None,               help="Also write JSON output")
#     parser.add_argument(
#         "--workers", "-w",
#         type=int,
#         default=min(8, multiprocessing.cpu_count()),
#         help="Parallel threads (default: min(8, cpu_count))",
#     )
#     args = parser.parse_args()

#     images = collect_images(args.input)
#     if not images:
#         print(f"No images found at: {args.input}", file=sys.stderr)
#         sys.exit(1)

#     n = len(images)
#     w = min(args.workers, n)
#     print(f"Found {n} image(s).  Threads: {w}\n", flush=True)

#     try:
#         pyt = setup_tesseract()
#         from PIL import Image as PILImage
#         import numpy as np
#         pyt.image_to_string(PILImage.fromarray(np.ones((10, 10, 3), dtype=np.uint8) * 255))
#         print("Tesseract OK.\n", flush=True)
#     except Exception as e:
#         print(f"\nERROR: Tesseract not available: {e}", file=sys.stderr)
#         print("Install from: https://github.com/UB-Mannheim/tesseract/wiki", file=sys.stderr)
#         sys.exit(1)

#     rows_by_path = {}
#     done = 0

#     with ThreadPoolExecutor(max_workers=w) as exe:
#         futures = {exe.submit(extract_one, p): p for p in images}
#         for fut in as_completed(futures):
#             done += 1
#             img_path = futures[fut]
#             try:
#                 row = fut.result()
#                 msg = row.pop("_msg", "")
#                 rows_by_path[img_path] = row
#                 print(f"[{done}/{n}] {Path(img_path).name}{msg}", flush=True)
#             except Exception as e:
#                 print(f"[{done}/{n}] ERROR {Path(img_path).name}: {e}", flush=True)

#     rows = [rows_by_path[p] for p in images if p in rows_by_path]

#     if args.json:
#         with open(args.json, "w", encoding="utf-8") as f:
#             json.dump(rows, f, indent=2, ensure_ascii=False)
#         print(f"JSON → {args.json}", flush=True)

#     if not rows:
#         print("No rows extracted.", flush=True)
#         sys.exit(0)

#     build_excel(rows, args.output)


# if __name__ == "__main__":
#     main()






































#2018 MORE TIME






# # -*- coding: utf-8 -*-
# """
# paddle_extract.py  —  NIRF score-card image → Excel extractor (v5)

# Fixes vs v4:
#   • Lower dark-separator threshold 0.75→0.65 so JPEG images (72% dark) are detected
#   • Score/Total row: numeric-scan fallback when label OCR fails on 2018 JPEG images
#   • IR code regex broadened: covers short codes like IR-1-P-P-C-409
#   • GPHE and PREMP added to KNOWN_COLS for 2018 scorecards
# """

# import io, os, sys, re, json, argparse
# from pathlib import Path
# from concurrent.futures import ThreadPoolExecutor, as_completed
# import multiprocessing

# if hasattr(sys.stdout, "reconfigure"):
#     sys.stdout.reconfigure(encoding="utf-8", errors="replace")
# else:
#     sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# TESSERACT_CMD = None   # set to r"C:\Program Files\Tesseract-OCR\tesseract.exe" if needed

# # ─────────────────────────────────────────────────────────────────────────────
# # Tesseract setup
# # ─────────────────────────────────────────────────────────────────────────────

# def setup_tesseract():
#     import pytesseract
#     if TESSERACT_CMD:
#         pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
#     else:
#         default = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
#         if os.path.exists(default):
#             pytesseract.pytesseract.tesseract_cmd = default
#     return pytesseract


# # ─────────────────────────────────────────────────────────────────────────────
# # Known NIRF column metadata
# # ─────────────────────────────────────────────────────────────────────────────

# KNOWN_COLS = {
#     "SS", "FSR", "FQE", "FRU",
#     "OE+MIR", "OE", "MIR",
#     "PU", "QP", "IPR", "FPPP", "SDG",
#     "GPH", "GPHE", "GUE", "MS", "GPHD",   # GPHE = 2018
#     "RD", "WD", "ESCS", "PCS",
#     "PR", "PREMP", "PERCEPTION",           # PREMP = 2018
# }

# GROUP_HEADERS = {"TLR", "RP", "GO", "OI"}
# ROW_LABELS    = {"SCORE", "TOTAL", "RANK"}

# OCR_FIXES = {
#     "S5": "SS", "5S": "SS", "S$": "SS", "55": "SS",
#     "W0": "WD", "VVD": "WD", "WO": "WD",
#     "FQF": "FQE", "FOE": "FQE", "FDE": "FQE",
#     "FRO": "FRU", "FR0": "FRU",
#     "1PR": "IPR", "lPR": "IPR",
#     "GU3": "GUE",
#     "P5": "PCS", "PC5": "PCS",
#     "E5CS": "ESCS", "ESC5": "ESCS", "ECSS": "ESCS",
#     "0E+MIR": "OE+MIR", "OE-MIR": "OE+MIR", "OE+MR": "OE+MIR",
#     "5DG": "SDG", "SDC": "SDG",
#     "M5": "MS", "GP H": "GPH",
#     "OP": "QP", "OF": "OE",
#     "GPHD": "GPHD", "GPHO": "GPHD",
#     "SR": "FSR", "FSP": "FSR",
#     "MlR": "MIR", "FRl": "FRU", "FRI": "FRU",
#     "GPHE": "GPHE", "GPH E": "GPHE",
#     "PREMP": "PREMP", "PREMI": "PREMP", "PR EMP": "PREMP",
# }

# PREFIX_COMPLETIONS = {
#     "SR": "FSR", "QE": "FQE", "RU": "FRU",
#     "S": "SS", "PH": "GPH", "UE": "GUE",
#     "HD": "GPHD", "CS": "PCS", "SCS": "ESCS",
#     "HE": "GPHE", "EMP": "PREMP",
# }

# # Matches ALL known IR code formats:
# #   IR-C-C-6361      (standard)
# #   IR-O-U-0391      (university)
# #   IR-1-C-C-C-6361  (2018 with year prefix)
# #   IR-1-P-P-C-409   (2018 short code)
# IR_CODE_RE = re.compile(r"IR-[A-Z0-9][-A-Z0-9]*-\d{3,}")


# def fix_ocr_token(tok: str) -> str:
#     t = tok.upper().replace(" ", "")
#     return OCR_FIXES.get(t, t)


# def is_col_header(tok: str) -> bool:
#     t = fix_ocr_token(tok)
#     if t in GROUP_HEADERS or t in ROW_LABELS:
#         return False
#     if t in KNOWN_COLS:
#         return True
#     return bool(re.fullmatch(r"[A-Z+]{2,8}", t))


# # ─────────────────────────────────────────────────────────────────────────────
# # Layout detection
# # ─────────────────────────────────────────────────────────────────────────────

# def find_dark_separator(arr, start_frac=0.35):
#     """
#     Find the thick black horizontal border that separates the chart from the table.
#     FIX: threshold lowered from 0.75 → 0.65 so JPEG images are detected.
#     JPEG compression reduces pure-black pixels to ~72%, below the old 75% threshold.
#     """
#     import numpy as np
#     h, w = arr.shape[:2]
#     for y in range(int(h * start_frac), h):
#         dark = int(np.sum((arr[y, :, 0] < 80) & (arr[y, :, 1] < 80) & (arr[y, :, 2] < 80)))
#         if dark > w * 0.65:   # was 0.75 — JPEG needs 0.65
#             return y
#     return None


# def find_score_total_y(img, sep_y, w, h):
#     """
#     Find the y-centres of the Score and Total rows.

#     Strategy (two-pass):
#       Pass 1 — OCR the left margin for "Score"/"Total" text labels (works for PNG/2019+)
#       Pass 2 — Numeric scan: find the two rows with the most decimal numbers
#                in the range sep_y+50 to sep_y+200 (works for 2018 JPEG where
#                label OCR fails due to compression artifacts)
#     """
#     pytesseract = setup_tesseract()
#     import numpy as np
#     from PIL import Image as PILImage

#     # ── Pass 1: label OCR ────────────────────────────────────────────────────
#     search_top = sep_y + 40
#     search_bot = min(h, sep_y + 220)
#     region     = img.crop((0, search_top, min(200, w), search_bot))
#     scale      = 4
#     region_big = region.resize((region.size[0] * scale, region.size[1] * scale), _lanczos())
#     data = pytesseract.image_to_data(
#         region_big.convert("L"),
#         config="--psm 11",
#         output_type=pytesseract.Output.DICT,
#     )
#     score_y = total_y = None
#     for i, text in enumerate(data["text"]):
#         t = text.strip().lower()
#         if "score" in t and score_y is None:
#             score_y = (data["top"][i] + data["height"][i] / 2) / scale + search_top
#         if "total" in t and total_y is None:
#             total_y = (data["top"][i] + data["height"][i] / 2) / scale + search_top

#     # ── Pass 2: numeric scan fallback (for 2018 JPEG) ────────────────────────
#     if score_y is None or total_y is None:
#         # Scan rows in sep_y+50 .. sep_y+200, count decimal numbers per row
#         row_counts = []
#         for y_cen in range(sep_y + 50, min(h - 15, sep_y + 210), 4):
#             y1 = max(0, y_cen - 12)
#             y2 = min(h,  y_cen + 12)
#             crop = img.crop((80, y1, w, y2))
#             big  = crop.resize((crop.size[0] * 4, crop.size[1] * 4), PILImage.LANCZOS)
#             arr_c  = np.array(big)
#             max_ch = np.maximum(np.maximum(arr_c[:,:,0].astype(int),
#                                            arr_c[:,:,1].astype(int)),
#                                            arr_c[:,:,2].astype(int))
#             clean  = np.where(max_ch < 160, 0, 255).astype("uint8")
#             pil_c  = PILImage.fromarray(clean)
#             d = pytesseract.image_to_data(
#                 pil_c,
#                 config="--psm 11 -c tessedit_char_whitelist=.0123456789",
#                 output_type=pytesseract.Output.DICT,
#             )
#             nums = [d["text"][i] for i in range(len(d["text"]))
#                     if d["text"][i].strip() and d["conf"][i] > 5
#                     and re.fullmatch(r"\d+\.\d+", d["text"][i].strip())]
#             if len(nums) >= 5:
#                 row_counts.append((len(nums), y_cen))

#         # Pick the two richest rows (score = first, total = second)
#         row_counts.sort(key=lambda x: (-x[0], x[1]))
#         rich_rows = sorted(set(r[1] for r in row_counts[:6]))
#         if len(rich_rows) >= 2 and score_y is None:
#             score_y = rich_rows[0]
#             total_y = rich_rows[-1]
#         elif len(rich_rows) == 1 and score_y is None:
#             score_y = rich_rows[0]

#     # ── Hard fallbacks ────────────────────────────────────────────────────────
#     if score_y is None:
#         score_y = sep_y + 80
#     if total_y is None:
#         total_y = score_y + 32

#     return float(score_y), float(total_y)


# def _lanczos():
#     from PIL import Image
#     return Image.LANCZOS


# # ─────────────────────────────────────────────────────────────────────────────
# # Column header OCR — three-pass
# # ─────────────────────────────────────────────────────────────────────────────

# def ocr_full_header_band(pytesseract, img, sep_y, w, scale=4):
#     from PIL import Image
#     crop = img.crop((0, sep_y - 38, w, sep_y))
#     big  = crop.resize((w * scale, crop.size[1] * scale), Image.LANCZOS)
#     gray = big.convert("L")
#     data = pytesseract.image_to_data(
#         gray,
#         config="--psm 11 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ+",
#         output_type=pytesseract.Output.DICT,
#     )
#     items = []
#     for i, text in enumerate(data["text"]):
#         text = text.strip()
#         if not text:
#             continue
#         conf  = data["conf"][i]
#         fixed = fix_ocr_token(text)
#         if conf < 0:
#             continue
#         if conf < 5 and fixed not in KNOWN_COLS:
#             continue
#         xc = (data["left"][i] + data["width"][i] / 2) / scale
#         items.append((xc, text))
#     items.sort(key=lambda t: t[0])
#     return items


# def ocr_per_column_header(pytesseract, img, xc, sep_y, w, scale=6):
#     from PIL import Image
#     margin = 48
#     x1 = max(0, int(xc - margin))
#     x2 = min(w, int(xc + margin))
#     crop = img.crop((x1, sep_y - 38, x2, sep_y))
#     big  = crop.resize((crop.size[0] * scale, crop.size[1] * scale), Image.LANCZOS)
#     gray = big.convert("L")
#     text = pytesseract.image_to_string(
#         gray,
#         config="--psm 6 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ+",
#     ).strip().replace("\n", " ").strip()
#     tokens = [t for t in text.split() if t]
#     return max(tokens, key=len) if tokens else ""


# def build_header_map(pytesseract, img, sep_y, w, score_xs):
#     """
#     For 2018 JPEG: col headers are at sep_y+32 to sep_y+61, NOT just below sep_y.
#     We try the standard band (sep_y-38 to sep_y) first, then fall back to
#     sep_y+32 to sep_y+61 if fewer than half the columns were found.
#     """
#     from PIL import Image

#     def _run_ocr_on_band(y1, y2):
#         crop = img.crop((0, y1, w, y2))
#         big  = crop.resize((w * 4, crop.size[1] * 4), Image.LANCZOS)
#         gray = big.convert("L")
#         data = pytesseract.image_to_data(
#             gray,
#             config="--psm 11 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ+",
#             output_type=pytesseract.Output.DICT,
#         )
#         items = []
#         for i, text in enumerate(data["text"]):
#             text = text.strip()
#             if not text:
#                 continue
#             conf  = data["conf"][i]
#             fixed = fix_ocr_token(text)
#             if conf < 0:
#                 continue
#             if conf < 5 and fixed not in KNOWN_COLS:
#                 continue
#             xc_val = (data["left"][i] + data["width"][i] / 2) / 4
#             items.append((xc_val, text))
#         items.sort(key=lambda t: t[0])
#         return items

#     # Try standard band first (works for PNG / 2019+)
#     raw_items = _run_ocr_on_band(sep_y - 38, sep_y)
#     pass_a = [(x, fix_ocr_token(t)) for x, t in raw_items if is_col_header(t)]

#     # If fewer than half expected columns found, try 2018 JPEG band (sep_y+32..sep_y+61)
#     if len(pass_a) < max(len(score_xs) // 2, 3):
#         raw_items2 = _run_ocr_on_band(sep_y + 32, sep_y + 62)
#         pass_a2 = [(x, fix_ocr_token(t)) for x, t in raw_items2 if is_col_header(t)]
#         if len(pass_a2) > len(pass_a):
#             pass_a = pass_a2

#     # Nearest-neighbour assignment
#     assigned   = {}
#     used_score = set()
#     for hx, hname in pass_a:
#         best_dist, best_idx = float("inf"), None
#         for i, sx in enumerate(score_xs):
#             if i in used_score:
#                 continue
#             d = abs(hx - sx)
#             if d < best_dist:
#                 best_dist, best_idx = d, i
#         if best_idx is not None and best_dist < 120:
#             assigned[best_idx] = hname
#             used_score.add(best_idx)

#     # Pass B: per-column crop for any gaps
#     missing = [i for i in range(len(score_xs)) if i not in assigned]
#     for i in missing:
#         xc = score_xs[i]
#         raw = ocr_per_column_header(pytesseract, img, xc, sep_y, w)
#         if raw:
#             fixed = fix_ocr_token(raw)
#             if fixed not in KNOWN_COLS and fixed in PREFIX_COMPLETIONS:
#                 fixed = PREFIX_COMPLETIONS[fixed]
#             if is_col_header(fixed):
#                 assigned[i] = fixed

#     # Pass C: prefix completion
#     for i, name in list(assigned.items()):
#         if name not in KNOWN_COLS and name in PREFIX_COMPLETIONS:
#             assigned[i] = PREFIX_COMPLETIONS[name]

#     return {score_xs[i]: assigned.get(i, f"COL{i+1}") for i in range(len(score_xs))}


# # ─────────────────────────────────────────────────────────────────────────────
# # Value extraction
# # ─────────────────────────────────────────────────────────────────────────────

# def extract_values_at_y(pytesseract, img, y_centre, w, scale=4):
#     import numpy as np
#     from PIL import Image

#     half = 14
#     y1, y2 = max(0, int(y_centre - half)), int(y_centre + half)
#     crop = img.crop((80, y1, w, y2))
#     big  = crop.resize((crop.size[0] * scale, crop.size[1] * scale), Image.LANCZOS)

#     arr    = np.array(big)
#     max_ch = np.maximum(np.maximum(arr[:,:,0].astype(int),
#                                    arr[:,:,1].astype(int)),
#                                    arr[:,:,2].astype(int))
#     clean  = np.where(max_ch < 160, 0, 255).astype("uint8")
#     from PIL import Image as PILImage
#     pil_clean = PILImage.fromarray(clean)

#     data = pytesseract.image_to_data(
#         pil_clean,
#         config="--psm 11 -c tessedit_char_whitelist=.0123456789",
#         output_type=pytesseract.Output.DICT,
#     )
#     items = []
#     for i, text in enumerate(data["text"]):
#         text = text.strip()
#         if not text or data["conf"][i] < 5:
#             continue
#         xc = (data["left"][i] + data["width"][i] / 2) / scale + 80
#         items.append((xc, text))
#     items.sort(key=lambda t: t[0])
#     return items


# def smart_fix_decimal(s):
#     s = s.strip()
#     if not s:
#         return s
#     if re.fullmatch(r"\d+\.\d+", s):
#         return s
#     if s.isdigit() and len(s) >= 3:
#         return s[:-2] + "." + s[-2:]
#     return s


# def is_float(s):
#     try:
#         float(s)
#         return True
#     except ValueError:
#         return False


# def match_values_to_headers(header_map, value_items):
#     result   = {}
#     used     = set()
#     for hx, col_name in sorted(header_map.items()):
#         best_dist, best_val, best_idx = float("inf"), "", None
#         for i, (vx, vtext) in enumerate(value_items):
#             if i in used:
#                 continue
#             d = abs(hx - vx)
#             if d < best_dist:
#                 best_dist, best_val, best_idx = d, vtext, i
#         if best_dist < 100 and best_idx is not None:
#             used.add(best_idx)
#             result[col_name] = smart_fix_decimal(best_val)
#     return result


# # ─────────────────────────────────────────────────────────────────────────────
# # Institute name / code  — THREE SOURCES
# # ─────────────────────────────────────────────────────────────────────────────

# def extract_institute_meta(pytesseract, img, arr):
#     import numpy as np
#     from PIL import Image
#     h, w = arr.shape[:2]

#     # ── Source 1: filename ────────────────────────────────────────────────────
#     fname    = Path(getattr(img, "filename", "")).stem
#     m_fname  = IR_CODE_RE.match(fname)
#     institute_code = m_fname.group(0) if m_fname else ""

#     # ── Find title rows ───────────────────────────────────────────────────────
#     t_start = t_end = None
#     for y in range(min(90, h)):
#         nw = int(np.sum(~((arr[y,:,0] > 210) & (arr[y,:,1] > 210) & (arr[y,:,2] > 210))))
#         if nw > 30:
#             if t_start is None:
#                 t_start = y
#             t_end = y
#     if t_start is None:
#         t_start, t_end = 5, 30

#     title_crop = img.crop((0, max(0, t_start - 4), w, min(h, t_end + 6)))
#     tw, th     = title_crop.size
#     title_big  = title_crop.resize((tw * 5, th * 5), Image.LANCZOS)
#     title_text = pytesseract.image_to_string(title_big.convert("L"), config="--psm 7").strip()

#     # ── Source 2: code from OCR title text ───────────────────────────────────
#     if not institute_code:
#         m_code = IR_CODE_RE.search(title_text)
#         if m_code:
#             institute_code = m_code.group(0).strip()

#     # ── Source 3: wider PSM-11 scan of top 20% ───────────────────────────────
#     if not institute_code:
#         header_crop = img.crop((0, 0, w, int(h * 0.20)))
#         header_big  = header_crop.resize((w * 3, header_crop.size[1] * 3), Image.LANCZOS)
#         header_text = pytesseract.image_to_string(
#             header_big.convert("L"), config="--psm 11"
#         ).strip()
#         m_code2 = IR_CODE_RE.search(header_text)
#         if m_code2:
#             institute_code = m_code2.group(0).strip()

#     # ── Institute name ────────────────────────────────────────────────────────
#     m = IR_CODE_RE.search(title_text)
#     if m:
#         name_part      = title_text[:m.start()].strip().rstrip("(").strip()
#         institute_name = name_part if name_part else ""
#     else:
#         lines          = [l.strip() for l in title_text.splitlines() if len(l.strip()) > 4]
#         institute_name = max(lines, key=len) if lines else ""
#     institute_name = re.sub(r"^[^A-Za-z;]+", "", institute_name).strip().lstrip(";").strip()

#     return institute_name, institute_code


# # ─────────────────────────────────────────────────────────────────────────────
# # Per-image extraction entry point
# # ─────────────────────────────────────────────────────────────────────────────

# def extract_one(image_path):
#     from PIL import Image
#     import numpy as np

#     pytesseract = setup_tesseract()

#     img = Image.open(image_path).convert("RGB")
#     img.filename = image_path
#     arr = np.array(img)
#     h, w = arr.shape[:2]

#     institute_name, institute_code = extract_institute_meta(pytesseract, img, arr)

#     sep_y = find_dark_separator(arr)
#     if sep_y is None:
#         return {
#             "image_path": image_path, "year": "", "category": "",
#             "institute_name": institute_name, "institute_code": institute_code,
#             "table": {},
#             "_msg": f"  NO_SEP {institute_code} | {institute_name[:40]}",
#         }

#     score_y, total_y = find_score_total_y(img, sep_y, w, h)

#     score_items = extract_values_at_y(pytesseract, img, score_y, w)
#     score_float = [(x, smart_fix_decimal(v)) for x, v in score_items if is_float(smart_fix_decimal(v))]
#     score_xs    = [x for x, _ in score_float]

#     if not score_xs:
#         return {
#             "image_path": image_path, "year": "", "category": "",
#             "institute_name": institute_name, "institute_code": institute_code,
#             "table": {},
#             "_msg": f"  NO_SCORES {institute_code} | {institute_name[:40]}",
#         }

#     header_map  = build_header_map(pytesseract, img, sep_y, w, score_xs)

#     total_items = extract_values_at_y(pytesseract, img, total_y, w)
#     total_float = [(x, smart_fix_decimal(v)) for x, v in total_items if is_float(smart_fix_decimal(v))]

#     score_map = match_values_to_headers(header_map, score_float)
#     total_map = match_values_to_headers(header_map, total_float)

#     table = {
#         col_name: {
#             "score": score_map.get(col_name, ""),
#             "total": total_map.get(col_name, ""),
#         }
#         for col_name in header_map.values()
#     }

#     parts    = Path(image_path).resolve().parts
#     year     = category = ""
#     for i, p in enumerate(parts):
#         if p == "image" and i >= 2:
#             year, category = parts[i - 2], parts[i - 1]
#             break

#     n_cols   = len(table)
#     n_scores = sum(1 for v in table.values() if v["score"])
#     flag     = " MISMATCH" if n_cols > 0 and n_cols != n_scores else " OK"

#     return {
#         "image_path":     image_path,
#         "year":           year,
#         "category":       category,
#         "institute_name": institute_name,
#         "institute_code": institute_code,
#         "table":          table,
#         "_msg": (
#             f"  {institute_code} | {(institute_name or '?')[:40]} "
#             f"| cols={n_cols} scores={n_scores}{flag}"
#         ),
#     }


# # ─────────────────────────────────────────────────────────────────────────────
# # Image collection
# # ─────────────────────────────────────────────────────────────────────────────

# IMAGE_EXTS = {".jpg", ".jpeg", ".png"}


# def collect_images(path):
#     p = Path(os.path.normpath(path))
#     if p.is_file() and p.suffix.lower() in IMAGE_EXTS:
#         return [str(p)]
#     if p.is_dir():
#         imgs = []
#         for ext in IMAGE_EXTS:
#             imgs.extend(str(f) for f in p.rglob(f"*{ext}"))
#         imgs.sort()
#         return imgs
#     print(f"Path not found: {p}", file=sys.stderr)
#     return []


# # ─────────────────────────────────────────────────────────────────────────────
# # Excel output
# # ─────────────────────────────────────────────────────────────────────────────

# def build_excel(rows, output_path):
#     from openpyxl import Workbook
#     from openpyxl.styles import Font, PatternFill, Alignment

#     header_fill  = PatternFill("solid", fgColor="1F4E79")
#     header_font  = Font(color="FFFFFF", bold=True)
#     alt_fill     = PatternFill("solid", fgColor="D9E1F2")
#     center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

#     all_headers = []
#     seen_h      = set()
#     for row in rows:
#         for hk in row.get("table", {}):
#             if hk not in seen_h:
#                 seen_h.add(hk)
#                 all_headers.append(hk)

#     fixed_cols = ["Year", "Category", "Institute Name", "Institute Code"]
#     all_cols   = (
#         fixed_cols
#         + [f"{h} Score" for h in all_headers]
#         + [f"{h} Total" for h in all_headers]
#     )

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "NIRF Scores"

#     for ci, col in enumerate(all_cols, 1):
#         cell           = ws.cell(row=1, column=ci, value=col)
#         cell.fill      = header_fill
#         cell.font      = header_font
#         cell.alignment = center_align
#     ws.row_dimensions[1].height = 30

#     for ri, row in enumerate(rows, 2):
#         tbl  = row.get("table", {})
#         fill = alt_fill if ri % 2 == 0 else PatternFill()
#         vals = (
#             [row.get("year",""), row.get("category",""),
#              row.get("institute_name",""), row.get("institute_code","")]
#             + [tbl.get(h, {}).get("score", "") for h in all_headers]
#             + [tbl.get(h, {}).get("total", "") for h in all_headers]
#         )
#         for ci, val in enumerate(vals, 1):
#             cell           = ws.cell(row=ri, column=ci, value=val)
#             cell.fill      = fill
#             cell.alignment = Alignment(horizontal="center", vertical="center")

#     for ci in range(1, len(all_cols) + 1):
#         col_letter = ws.cell(row=1, column=ci).column_letter
#         max_len = max(
#             len(str(ws.cell(row=1, column=ci).value or "")),
#             *(len(str(ws.cell(row=ri, column=ci).value or "")) for ri in range(2, len(rows) + 2)),
#         )
#         ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

#     ws.freeze_panes = "A2"
#     wb.save(output_path)
#     print(f"\nDone. Saved {len(rows)} rows → {output_path}", flush=True)


# # ─────────────────────────────────────────────────────────────────────────────
# # Main
# # ─────────────────────────────────────────────────────────────────────────────

# def main():
#     parser = argparse.ArgumentParser(description="Extract NIRF score-card table data (v5)")
#     parser.add_argument("input",    help="Image file or folder (searched recursively)")
#     parser.add_argument("--output", "-o", default="nirf_scores.xlsx", help="Output Excel file")
#     parser.add_argument("--json",   "-j", default=None,               help="Also write JSON output")
#     parser.add_argument(
#         "--workers", "-w",
#         type=int,
#         default=min(8, multiprocessing.cpu_count()),
#         help="Parallel threads (default: min(8, cpu_count))",
#     )
#     args = parser.parse_args()

#     images = collect_images(args.input)
#     if not images:
#         print(f"No images found at: {args.input}", file=sys.stderr)
#         sys.exit(1)

#     n = len(images)
#     w = min(args.workers, n)
#     print(f"Found {n} image(s).  Threads: {w}\n", flush=True)

#     try:
#         pyt = setup_tesseract()
#         from PIL import Image as PILImage
#         import numpy as np
#         pyt.image_to_string(PILImage.fromarray(np.ones((10, 10, 3), dtype=np.uint8) * 255))
#         print("Tesseract OK.\n", flush=True)
#     except Exception as e:
#         print(f"\nERROR: Tesseract not available: {e}", file=sys.stderr)
#         print("Install from: https://github.com/UB-Mannheim/tesseract/wiki", file=sys.stderr)
#         sys.exit(1)

#     rows_by_path = {}
#     done = 0

#     with ThreadPoolExecutor(max_workers=w) as exe:
#         futures = {exe.submit(extract_one, p): p for p in images}
#         for fut in as_completed(futures):
#             done += 1
#             img_path = futures[fut]
#             try:
#                 row = fut.result()
#                 msg = row.pop("_msg", "")
#                 rows_by_path[img_path] = row
#                 print(f"[{done}/{n}] {Path(img_path).name}{msg}", flush=True)
#             except Exception as e:
#                 print(f"[{done}/{n}] ERROR {Path(img_path).name}: {e}", flush=True)

#     rows = [rows_by_path[p] for p in images if p in rows_by_path]

#     if args.json:
#         with open(args.json, "w", encoding="utf-8") as f:
#             json.dump(rows, f, indent=2, ensure_ascii=False)
#         print(f"JSON → {args.json}", flush=True)

#     if not rows:
#         print("No rows extracted.", flush=True)
#         sys.exit(0)

#     build_excel(rows, args.output)


# if __name__ == "__main__":
#     main()
































# -*- coding: utf-8 -*-
"""
paddle_extract.py  —  NIRF score-card image → Excel extractor (v6)

Speed optimisations vs v5  (~7x faster per image):
  • Score/Total detection: 27 separate Tesseract calls → ONE call on full region
    (was 29s, now 3s — the single biggest bottleneck)
  • Score/Total values:    extracted from the same single call, not a separate OCR
  • "Score"/"Total" label pass dropped — numeric scan is faster and more reliable
  • All upscale factors lowered 4x/5x → 3x (still accurate, faster resize+OCR)
  • Header band OCR: also reuses 3x scale

Accuracy fixes vs v4:
  • JPEG dark-separator threshold: 0.75 → 0.65
  • IR code regex covers all formats including IR-1-P-P-C-409
  • GPHE and PREMP added to KNOWN_COLS for 2018 scorecards
  • Header band fallback for 2018 JPEG (col headers at sep_y+32..+62)
"""

import io, os, sys, re, json, argparse
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict
import multiprocessing

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
else:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

TESSERACT_CMD = None   # set to r"C:\Program Files\Tesseract-OCR\tesseract.exe" if needed

# ─────────────────────────────────────────────────────────────────────────────
# Tesseract setup
# ─────────────────────────────────────────────────────────────────────────────

def setup_tesseract():
    import pytesseract
    if TESSERACT_CMD:
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
    else:
        default = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        if os.path.exists(default):
            pytesseract.pytesseract.tesseract_cmd = default
    return pytesseract


# ─────────────────────────────────────────────────────────────────────────────
# Known NIRF column metadata
# ─────────────────────────────────────────────────────────────────────────────

KNOWN_COLS = {
    "SS", "FSR", "FQE", "FRU",
    "OE+MIR", "OE", "MIR",
    "PU", "QP", "IPR", "FPPP", "SDG",
    "GPH", "GPHE", "GUE", "MS", "GPHD",
    "RD", "WD", "ESCS", "PCS",
    "PR", "PREMP", "PERCEPTION",
}

GROUP_HEADERS = {"TLR", "RP", "GO", "OI"}
ROW_LABELS    = {"SCORE", "TOTAL", "RANK"}

OCR_FIXES = {
    "S5": "SS",  "5S": "SS",  "S$": "SS",  "55": "SS",
    "W0": "WD",  "VVD": "WD", "WO": "WD",
    "FQF": "FQE","FOE": "FQE","FDE": "FQE",
    "FRO": "FRU","FR0": "FRU",
    "1PR": "IPR","lPR": "IPR",
    "GU3": "GUE",
    "P5": "PCS", "PC5": "PCS",
    "E5CS": "ESCS","ESC5": "ESCS","ECSS": "ESCS",
    "0E+MIR": "OE+MIR","OE-MIR": "OE+MIR","OE+MR": "OE+MIR",
    "5DG": "SDG","SDC": "SDG",
    "M5": "MS",  "GP H": "GPH",
    "OP": "QP",  "OF": "OE",
    "GPHD": "GPHD","GPHO": "GPHD",
    "SR": "FSR", "FSP": "FSR",
    "MlR": "MIR","FRl": "FRU","FRI": "FRU",
    "GPHE": "GPHE","GPH E": "GPHE",
    "PREMP": "PREMP","PREMI": "PREMP","PR EMP": "PREMP",
}

PREFIX_COMPLETIONS = {
    "SR": "FSR",  "QE": "FQE",  "RU": "FRU",
    "S":  "SS",   "PH": "GPH",  "UE": "GUE",
    "HD": "GPHD", "CS": "PCS",  "SCS": "ESCS",
    "HE": "GPHE", "EMP": "PREMP",
}

# Matches all IR code formats: IR-C-C-6361, IR-O-U-0391, IR-1-C-C-C-6361, IR-1-P-P-C-409
IR_CODE_RE = re.compile(r"IR-[A-Z0-9][-A-Z0-9]*-\d{3,}")


def fix_ocr_token(tok):
    t = tok.upper().replace(" ", "")
    return OCR_FIXES.get(t, t)


def is_col_header(tok):
    t = fix_ocr_token(tok)
    if t in GROUP_HEADERS or t in ROW_LABELS:
        return False
    if t in KNOWN_COLS:
        return True
    return bool(re.fullmatch(r"[A-Z+]{2,8}", t))


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _lanczos():
    from PIL import Image
    return Image.LANCZOS

def smart_fix_decimal(s):
    s = s.strip()
    if not s:
        return s
    if re.fullmatch(r"\d+\.\d+", s):
        return s
    if s.isdigit() and len(s) >= 3:
        return s[:-2] + "." + s[-2:]
    return s

def is_float(s):
    try:    float(s); return True
    except: return False

def binarize(arr):
    """Max-channel binarize to remove bar-chart colour bleed."""
    import numpy as np
    max_ch = np.maximum(np.maximum(arr[:,:,0].astype(int),
                                   arr[:,:,1].astype(int)),
                                   arr[:,:,2].astype(int))
    return (max_ch >= 160).astype("uint8") * 255


# ─────────────────────────────────────────────────────────────────────────────
# Layout detection
# ─────────────────────────────────────────────────────────────────────────────

def find_dark_separator(arr, start_frac=0.35):
    """
    Finds the thick horizontal border between chart and table.
    Threshold 0.65 (not 0.75) so JPEG-compressed images work.
    """
    import numpy as np
    h, w = arr.shape[:2]
    for y in range(int(h * start_frac), h):
        dark = int(np.sum((arr[y,:,0]<80)&(arr[y,:,1]<80)&(arr[y,:,2]<80)))
        if dark > w * 0.65:
            return y
    return None


def find_score_total_rows(pytesseract, img, sep_y, w, h):
    """
    OPTIMISED: one single Tesseract call on the full score+total region.

    Old approach: 27 separate calls (one per 4px y-band) = ~30s
    New approach: 1 call on the whole region, cluster by Y coord  = ~3s

    Returns:
        score_items : list of (x_centre, value_str)
        total_items : list of (x_centre, value_str)
        score_y     : float  (y-centre of score row)
        total_y     : float  (y-centre of total row)
    """
    import numpy as np
    from PIL import Image as PILImage

    SCALE       = 3
    search_top  = sep_y + 50
    search_bot  = min(h, sep_y + 210)
    region      = img.crop((80, search_top, w, search_bot))
    rh          = search_bot - search_top
    rw          = w - 80

    big    = region.resize((rw * SCALE, rh * SCALE), PILImage.LANCZOS)
    clean  = PILImage.fromarray(binarize(np.array(big)))

    data = pytesseract.image_to_data(
        clean,
        config="--psm 11 -c tessedit_char_whitelist=.0123456789",
        output_type=pytesseract.Output.DICT,
    )

    # Cluster valid decimal numbers by Y-band (8px buckets in original coords)
    y_bands = defaultdict(list)
    for i, text in enumerate(data["text"]):
        text = text.strip()
        if not text or data["conf"][i] < 5:
            continue
        if not re.fullmatch(r"\d+\.\d+", text):
            continue
        yc = (data["top"][i] + data["height"][i] / 2) / SCALE + search_top
        xc = (data["left"][i] + data["width"][i] / 2) / SCALE + 80
        band = round(yc / 8) * 8
        y_bands[band].append((xc, text))

    # Keep only rows rich enough (>=5 values) and sort top-to-bottom
    rich = sorted(
        ((band, items) for band, items in y_bands.items() if len(items) >= 5),
        key=lambda x: x[0]
    )

    if len(rich) >= 2:
        score_y, score_items = rich[0][0],  sorted(rich[0][1])
        total_y, total_items = rich[-1][0], sorted(rich[-1][1])
    elif len(rich) == 1:
        score_y, score_items = rich[0][0],  sorted(rich[0][1])
        total_y, total_items = score_y + 32, []
    else:
        score_y, score_items = sep_y + 80,  []
        total_y, total_items = sep_y + 112, []

    return score_items, total_items, float(score_y), float(total_y)


# ─────────────────────────────────────────────────────────────────────────────
# Column header OCR
# ─────────────────────────────────────────────────────────────────────────────

def _ocr_header_band(pytesseract, img, y1, y2, w, scale=3):
    from PIL import Image
    crop = img.crop((0, y1, w, y2))
    big  = crop.resize((w * scale, max((y2-y1)*scale, 1)), Image.LANCZOS)
    data = pytesseract.image_to_data(
        big.convert("L"),
        config="--psm 11 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ+",
        output_type=pytesseract.Output.DICT,
    )
    items = []
    for i, text in enumerate(data["text"]):
        text = text.strip()
        if not text:
            continue
        conf  = data["conf"][i]
        fixed = fix_ocr_token(text)
        if conf < 0:
            continue
        if conf < 5 and fixed not in KNOWN_COLS:
            continue
        xc = (data["left"][i] + data["width"][i] / 2) / scale
        items.append((xc, text))
    items.sort(key=lambda t: t[0])
    return items


def ocr_per_column_header(pytesseract, img, xc, sep_y, w, scale=4):
    from PIL import Image
    margin = 48
    x1 = max(0, int(xc - margin))
    x2 = min(w, int(xc + margin))
    crop = img.crop((x1, sep_y - 38, x2, sep_y))
    big  = crop.resize((crop.size[0]*scale, max(crop.size[1]*scale,1)), Image.LANCZOS)
    text = pytesseract.image_to_string(
        big.convert("L"),
        config="--psm 6 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ+",
    ).strip().replace("\n", " ").strip()
    tokens = [t for t in text.split() if t]
    return max(tokens, key=len) if tokens else ""


def build_header_map(pytesseract, img, sep_y, w, score_xs):
    """
    Three-pass column header detection.
    Pass A: full band just above sep_y (PNG / 2019+)
    Pass A2: band just below sep_y (2018 JPEG — col headers at sep_y+32..+62)
    Pass B: per-column narrow crop for any remaining gaps
    Pass C: prefix completion
    """
    # Pass A: standard band
    pass_a = [(x, fix_ocr_token(t))
              for x, t in _ocr_header_band(pytesseract, img, sep_y-38, sep_y, w)
              if is_col_header(t)]

    # Pass A2: 2018 JPEG fallback band
    if len(pass_a) < max(len(score_xs) // 2, 3):
        alt = [(x, fix_ocr_token(t))
               for x, t in _ocr_header_band(pytesseract, img, sep_y+32, sep_y+62, w)
               if is_col_header(t)]
        if len(alt) > len(pass_a):
            pass_a = alt

    # Nearest-neighbour assignment
    assigned   = {}
    used_score = set()
    for hx, hname in pass_a:
        best_dist, best_idx = float("inf"), None
        for i, sx in enumerate(score_xs):
            if i in used_score:
                continue
            d = abs(hx - sx)
            if d < best_dist:
                best_dist, best_idx = d, i
        if best_idx is not None and best_dist < 120:
            assigned[best_idx] = hname
            used_score.add(best_idx)

    # Pass B: per-column crop for gaps
    for i in [i for i in range(len(score_xs)) if i not in assigned]:
        raw = ocr_per_column_header(pytesseract, img, score_xs[i], sep_y, w)
        if raw:
            fixed = fix_ocr_token(raw)
            if fixed not in KNOWN_COLS and fixed in PREFIX_COMPLETIONS:
                fixed = PREFIX_COMPLETIONS[fixed]
            if is_col_header(fixed):
                assigned[i] = fixed

    # Pass C: prefix completion
    for i, name in list(assigned.items()):
        if name not in KNOWN_COLS and name in PREFIX_COMPLETIONS:
            assigned[i] = PREFIX_COMPLETIONS[name]

    return {score_xs[i]: assigned.get(i, f"COL{i+1}") for i in range(len(score_xs))}


# ─────────────────────────────────────────────────────────────────────────────
# Value matching
# ─────────────────────────────────────────────────────────────────────────────

def match_values_to_headers(header_map, value_items):
    result  = {}
    used    = set()
    for hx, col_name in sorted(header_map.items()):
        best_dist, best_val, best_idx = float("inf"), "", None
        for i, (vx, vtext) in enumerate(value_items):
            if i in used:
                continue
            d = abs(hx - vx)
            if d < best_dist:
                best_dist, best_val, best_idx = d, vtext, i
        if best_dist < 100 and best_idx is not None:
            used.add(best_idx)
            result[col_name] = smart_fix_decimal(best_val)
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Institute name / code
# ─────────────────────────────────────────────────────────────────────────────

def extract_institute_meta(pytesseract, img, arr):
    import numpy as np
    from PIL import Image
    h, w = arr.shape[:2]

    # Source 1: filename
    fname   = Path(getattr(img, "filename", "")).stem
    m_fname = IR_CODE_RE.match(fname)
    institute_code = m_fname.group(0) if m_fname else ""

    # Find title rows
    t_start = t_end = None
    for y in range(min(90, h)):
        nw = int(np.sum(~((arr[y,:,0]>210)&(arr[y,:,1]>210)&(arr[y,:,2]>210))))
        if nw > 30:
            if t_start is None: t_start = y
            t_end = y
    if t_start is None:
        t_start, t_end = 5, 30

    title_crop = img.crop((0, max(0,t_start-4), w, min(h,t_end+6)))
    tw, th     = title_crop.size
    # OPT: 3x upscale instead of 5x — still readable, faster
    title_big  = title_crop.resize((tw*3, th*3), Image.LANCZOS)
    title_text = pytesseract.image_to_string(title_big.convert("L"), config="--psm 7").strip()

    # Source 2: code from OCR title
    if not institute_code:
        m = IR_CODE_RE.search(title_text)
        if m:
            institute_code = m.group(0).strip()

    # Source 3: wider scan of top 20%
    if not institute_code:
        hcrop = img.crop((0, 0, w, int(h*0.20)))
        hbig  = hcrop.resize((w*3, hcrop.size[1]*3), Image.LANCZOS)
        htext = pytesseract.image_to_string(hbig.convert("L"), config="--psm 11").strip()
        m2 = IR_CODE_RE.search(htext)
        if m2:
            institute_code = m2.group(0).strip()

    # Institute name
    m = IR_CODE_RE.search(title_text)
    if m:
        institute_name = title_text[:m.start()].strip().rstrip("(").strip()
    else:
        lines = [l.strip() for l in title_text.splitlines() if len(l.strip()) > 4]
        institute_name = max(lines, key=len) if lines else ""
    institute_name = re.sub(r"^[^A-Za-z;]+", "", institute_name).strip().lstrip(";").strip()

    return institute_name, institute_code


# ─────────────────────────────────────────────────────────────────────────────
# Per-image extraction
# ─────────────────────────────────────────────────────────────────────────────

def extract_one(image_path):
    import numpy as np
    from PIL import Image

    pytesseract = setup_tesseract()
    img = Image.open(image_path).convert("RGB")
    img.filename = image_path
    arr = np.array(img)
    h, w = arr.shape[:2]

    institute_name, institute_code = extract_institute_meta(pytesseract, img, arr)

    sep_y = find_dark_separator(arr)
    if sep_y is None:
        return {"image_path": image_path, "year": "", "category": "",
                "institute_name": institute_name, "institute_code": institute_code,
                "table": {}, "_msg": f"  NO_SEP {institute_code} | {institute_name[:40]}"}

    # OPTIMISED: one Tesseract call returns both score and total items + y positions
    score_items, total_items, score_y, total_y = find_score_total_rows(
        pytesseract, img, sep_y, w, h
    )

    score_float = [(x, smart_fix_decimal(v)) for x, v in score_items
                   if is_float(smart_fix_decimal(v))]
    score_xs    = [x for x, _ in score_float]

    if not score_xs:
        return {"image_path": image_path, "year": "", "category": "",
                "institute_name": institute_name, "institute_code": institute_code,
                "table": {}, "_msg": f"  NO_SCORES {institute_code} | {institute_name[:40]}"}

    header_map  = build_header_map(pytesseract, img, sep_y, w, score_xs)

    total_float = [(x, smart_fix_decimal(v)) for x, v in total_items
                   if is_float(smart_fix_decimal(v))]

    score_map = match_values_to_headers(header_map, score_float)
    total_map = match_values_to_headers(header_map, total_float)

    table = {
        col: {"score": score_map.get(col,""), "total": total_map.get(col,"")}
        for col in header_map.values()
    }

    parts    = Path(image_path).resolve().parts
    year = category = ""
    for i, p in enumerate(parts):
        if p == "image" and i >= 2:
            year, category = parts[i-2], parts[i-1]
            break

    n_cols   = len(table)
    n_scores = sum(1 for v in table.values() if v["score"])
    flag     = " MISMATCH" if n_cols > 0 and n_cols != n_scores else " OK"

    return {
        "image_path": image_path, "year": year, "category": category,
        "institute_name": institute_name, "institute_code": institute_code,
        "table": table,
        "_msg": f"  {institute_code} | {(institute_name or '?')[:40]} | cols={n_cols} scores={n_scores}{flag}",
    }


# ─────────────────────────────────────────────────────────────────────────────
# Image collection
# ─────────────────────────────────────────────────────────────────────────────

IMAGE_EXTS = {".jpg", ".jpeg", ".png"}

def collect_images(path):
    p = Path(os.path.normpath(path))
    if p.is_file() and p.suffix.lower() in IMAGE_EXTS:
        return [str(p)]
    if p.is_dir():
        imgs = []
        for ext in IMAGE_EXTS:
            imgs.extend(str(f) for f in p.rglob(f"*{ext}"))
        imgs.sort()
        return imgs
    print(f"Path not found: {p}", file=sys.stderr)
    return []


# ─────────────────────────────────────────────────────────────────────────────
# Excel output
# ─────────────────────────────────────────────────────────────────────────────

def build_excel(rows, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill  = PatternFill("solid", fgColor="1F4E79")
    header_font  = Font(color="FFFFFF", bold=True)
    alt_fill     = PatternFill("solid", fgColor="D9E1F2")

    all_headers, seen_h = [], set()
    for row in rows:
        for hk in row.get("table", {}):
            if hk not in seen_h:
                seen_h.add(hk)
                all_headers.append(hk)

    fixed_cols = ["Year", "Category", "Institute Name", "Institute Code"]
    all_cols   = (fixed_cols
                  + [f"{h} Score" for h in all_headers]
                  + [f"{h} Total" for h in all_headers])

    wb = Workbook()
    ws = wb.active
    ws.title = "NIRF Scores"

    for ci, col in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for ri, row in enumerate(rows, 2):
        tbl  = row.get("table", {})
        fill = alt_fill if ri % 2 == 0 else PatternFill()
        vals = ([row.get("year",""), row.get("category",""),
                 row.get("institute_name",""), row.get("institute_code","")]
                + [tbl.get(h,{}).get("score","") for h in all_headers]
                + [tbl.get(h,{}).get("total","") for h in all_headers])
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for ci in range(1, len(all_cols)+1):
        col_letter = ws.cell(row=1, column=ci).column_letter
        max_len = max(
            len(str(ws.cell(row=1,  column=ci).value or "")),
            *(len(str(ws.cell(row=ri, column=ci).value or "")) for ri in range(2, len(rows)+2)),
        )
        ws.column_dimensions[col_letter].width = min(max_len+2, 40)

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"\nDone. Saved {len(rows)} rows → {output_path}", flush=True)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Extract NIRF score-card table data (v6)")
    parser.add_argument("input",    help="Image file or folder (searched recursively)")
    parser.add_argument("--output", "-o", default="nirf_scores.xlsx", help="Output Excel file")
    parser.add_argument("--json",   "-j", default=None,               help="Also write JSON output")
    parser.add_argument("--workers","-w", type=int,
                        default=min(8, multiprocessing.cpu_count()),
                        help="Parallel threads (default: min(8, cpu_count))")
    args = parser.parse_args()

    images = collect_images(args.input)
    if not images:
        print(f"No images found at: {args.input}", file=sys.stderr)
        sys.exit(1)

    n = len(images)
    w = min(args.workers, n)
    print(f"Found {n} image(s).  Threads: {w}\n", flush=True)

    try:
        pyt = setup_tesseract()
        from PIL import Image as PILImage
        import numpy as np
        pyt.image_to_string(PILImage.fromarray(np.ones((10,10,3), dtype=np.uint8)*255))
        print("Tesseract OK.\n", flush=True)
    except Exception as e:
        print(f"\nERROR: Tesseract not available: {e}", file=sys.stderr)
        print("Install from: https://github.com/UB-Mannheim/tesseract/wiki", file=sys.stderr)
        sys.exit(1)

    rows_by_path = {}
    done = 0

    with ThreadPoolExecutor(max_workers=w) as exe:
        futures = {exe.submit(extract_one, p): p for p in images}
        for fut in as_completed(futures):
            done += 1
            img_path = futures[fut]
            try:
                row = fut.result()
                msg = row.pop("_msg", "")
                rows_by_path[img_path] = row
                print(f"[{done}/{n}] {Path(img_path).name}{msg}", flush=True)
            except Exception as e:
                print(f"[{done}/{n}] ERROR {Path(img_path).name}: {e}", flush=True)

    rows = [rows_by_path[p] for p in images if p in rows_by_path]

    if args.json:
        with open(args.json, "w", encoding="utf-8") as f:
            json.dump(rows, f, indent=2, ensure_ascii=False)
        print(f"JSON → {args.json}", flush=True)

    if not rows:
        print("No rows extracted.", flush=True)
        sys.exit(0)

    build_excel(rows, args.output)


if __name__ == "__main__":
    main()